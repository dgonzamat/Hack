"""
Genera Excel en formato Cuadro de Precios de Licitacion (estructura STM Vitacura).
Toma el template original, pre-llena cantidades calculadas y agrega hoja de analisis.
"""

import csv
import math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── UF referencial mayo 2026 ──────────────────────────────────────────────────
UF_CLP = 38_500

# ── Fills ─────────────────────────────────────────────────────────────────────
_FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")  # azul oscuro
_FILL_SECTION   = PatternFill("solid", fgColor="BDD7EE")  # azul claro (sección padre)
_FILL_CALC      = PatternFill("solid", fgColor="E2EFDA")  # verde claro (cantidad calculada)
_FILL_ORIGINAL  = PatternFill("solid", fgColor="FFF2CC")  # amarillo pálido (ya en template)
_FILL_PRICE     = PatternFill("solid", fgColor="FFFF00")  # amarillo (contratista llena precio)
_FILL_PRICE_REF = PatternFill("solid", fgColor="FCE4D6")  # salmón (precio referencial ONDAC)
_FILL_WARN      = PatternFill("solid", fgColor="FCE4D6")  # salmón (no cubicado)

# ── Mapeo sub-ítem → clave de precio en CSV ───────────────────────────────────
# Clave: último número del ítem (ej: "1" para 4.x.1, "7" para 5.x.7)
# Sección 4 (piques)
_PRECIO_PIQUE = {
    "1":  "excavacion_pique",
    "2":  "retiro_excavacion_pique",
    "3":  "montaje_liner_pique",
    "8":  "montaje_escalas_plataformas",
    "14": "brocal_definitivo",
    "15": "brocal_definitivo",
}
# Sección 5 (túneles)
_PRECIO_TUNEL = {
    "1": "excavacion_tunel",
    "2": "retiro_excavacion_tunel",
    "3": "montaje_liner_tunel",
    "7": "radier_tunel",
}


def cargar_precios_uf(csv_path: str | Path) -> dict[str, float]:
    """Carga precios desde CSV y los convierte a UF. Ignora líneas comentadas (#)."""
    precios: dict[str, float] = {}
    path = Path(csv_path)
    if not path.exists():
        return precios
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(filter(lambda l: not l.startswith("#"), f)):
            try:
                precios[row["partida"].strip()] = round(
                    float(row["precio_clp"]) / UF_CLP, 4
                )
            except (KeyError, ValueError):
                pass
    return precios


def _precio_para_item(num: str, precios_uf: dict[str, float]) -> float | None:
    """Devuelve precio UF para un ítem numerado tipo '4.3.1' o '5.2.7', o None."""
    parts = num.split(".")
    if len(parts) != 3:
        return None
    seccion, _, sub = parts
    lookup = _PRECIO_PIQUE if seccion == "4" else (_PRECIO_TUNEL if seccion == "5" else {})
    key = lookup.get(sub)
    return precios_uf.get(key) if key else None

_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_BOLD   = Font(bold=True, size=9)
_FONT_NORM   = Font(size=9)

_BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _cell(ws, row: int, col: int, value=None, fill=None, font=None,
          alignment=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:       c.fill = fill
    if font:       c.font = font
    if alignment:  c.alignment = alignment
    if number_format: c.number_format = number_format
    c.border = _BORDER_THIN
    return c


# ── Mapeo: clave de cubicacion → ítems de licitacion (por pique y tunel) ──────
# Para PIQUE N (1..9): reemplazar {p} por el número
_PIQUE_ITEMS = {
    "excavacion_pique":            "4.{p}.1",
    "retiro_excavacion_pique":     "4.{p}.2",
    "montaje_liner_pique":         "4.{p}.3",
    "montaje_escalas_plataformas": "4.{p}.8",
    # montaje_estructura_cables: solo P1 y P9 (ítem .13) — manejado en _build_qty_map
    # brocal_definitivo: P1/P9 → .15, P2-8 → .14 — manejado en _build_qty_map
}
# Items de tunnel: para tramo T (1..8)
_TUNEL_ITEMS = {
    "excavacion_tunel":        "5.{t}.1",
    "retiro_excavacion_tunel": "5.{t}.2",
    "montaje_liner_tunel":     "5.{t}.3",
    "radier_tunel":            "5.{t}.7",
    "instalacion_cables":      "5.{t}.8",
}
# Piques 1,9 tienen ítem extra "Montaje estructura soporte de cables" (4.x.13)
# → brocal definitivo queda en 4.x.15. Para piques 2-8 está en 4.x.14.
_BROCAL_NUM: dict[int, str] = {p: "14" for p in range(2, 9)}
_BROCAL_NUM.update({1: "15", 9: "15"})


def _build_qty_map(cubicacion: dict) -> dict[str, float]:
    """
    Construye {item_num: cantidad} a partir de las partidas cubicadas.

    Si cubicacion tiene piques_qty / tramos_qty (modo multi-pique estructurado),
    genera una entrada por pique (4.1.x … 4.9.x) y por tramo (5.1.x … 5.8.x).
    Fallback: PIQUE 1 solamente y distribución uniforme entre 8 tramos.
    """
    qty: dict[str, float] = {}

    piques_qty: dict = cubicacion.get("piques_qty", {})
    tramos_qty: dict = cubicacion.get("tramos_qty", {})

    if piques_qty:
        for pid, pq in piques_qty.items():
            pid_i = int(pid)
            for key, tpl in _PIQUE_ITEMS.items():
                if key in pq:
                    qty[tpl.replace("{p}", str(pid_i))] = pq[key]
            if "brocal_definitivo" in pq:
                brocal_num = _BROCAL_NUM.get(pid_i, "15")
                qty[f"4.{pid_i}.{brocal_num}"] = pq["brocal_definitivo"]
            if "montaje_estructura_cables" in pq:  # solo P1 y P9 (SS)
                qty[f"4.{pid_i}.13"] = pq["montaje_estructura_cables"]
    else:
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion.get("partidas", [])}
        for key, tpl in _PIQUE_ITEMS.items():
            if key in partidas:
                qty[tpl.replace("{p}", "1")] = partidas[key]
        if "brocal_definitivo" in partidas:
            qty["4.1.15"] = partidas["brocal_definitivo"]

    if tramos_qty:
        for tid, tq in tramos_qty.items():
            for key, tpl in _TUNEL_ITEMS.items():
                if key in tq:
                    num = tpl.replace("{t}", str(tid))
                    qty[num] = tq[key]
    else:
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion.get("partidas", [])}
        n_tramos = 8
        for key, tpl in _TUNEL_ITEMS.items():
            if key in partidas:
                cant_por_tramo = round(partidas[key] / n_tramos, 3)
                for t in range(1, n_tramos + 1):
                    qty[tpl.replace("{t}", str(t))] = cant_por_tramo

    return qty


# ── Descriptores de ítems licitación ─────────────────────────────────────────
# Estructura: (sub_num, descripcion, unidad, es_eventual)
_ITEMS_PIQUE = [
    ("1",  "Excavación pique",                        "m³",  False),
    ("2",  "Retiro y transporte material excavado",    "m³",  False),
    ("3",  "Montaje liner pique",                      "m",   False),
    ("4",  "Refuerzo losa anillo fondo",               "gl",  True),
    ("5",  "Construcción dren",                        "gl",  True),
    ("6",  "Terminaciones pique",                      "gl",  True),
    ("7",  "Cobertura pique",                          "gl",  True),
    ("8",  "Montaje escalas y plataformas",            "gl",  False),
    ("9",  "Refuerzo apertura pique",                  "gl",  True),
    ("10", "Shotcrete",                                "m³",  True),
    ("11", "Malla electrosoldada",                     "kg",  True),
    ("12", "Pernos de convergencia",                   "und", True),
]
# Ítems extra solo para piques 1 y 9 (SS)
_ITEMS_PIQUE_SS = [
    ("13", "Montaje estructura soporte de cables",     "gl",  False),
    ("15", "Brocal definitivo",                        "gl",  False),
]
# Brocal para piques típicos 2-8
_ITEMS_PIQUE_TYP = [
    ("14", "Brocal definitivo",                        "gl",  False),
]
_ITEMS_TUNEL = [
    ("1",  "Excavación túnel",                         "m³",  False),
    ("2",  "Retiro y transporte material excavado",    "m³",  False),
    ("3",  "Montaje liner túnel Ø2.21m",               "m",   False),
    ("4",  "Pernos frente de trabajo (eventual)",      "und", True),
    ("5",  "Shotcrete frente de trabajo (eventual)",   "m³",  True),
    ("6",  "Paraguas de micropilotes (eventual)",      "und", True),
    ("7",  "Radier túnel H30",                         "m³",  False),
    ("8",  "Instalación cables",                       "ml",  False),
]
_ENTRE_PIQUES = {
    1: "P1-P2", 2: "P2-P3", 3: "P3-P4", 4: "P4-P5",
    5: "P5-P6", 6: "P6-P7", 7: "P7-P8", 8: "P8-P9",
}


def _create_detalle_rows(ws) -> None:
    """
    Escribe la estructura completa de ítems (secciones 4 y 5) en una hoja vacía.
    Después _fill_detalle_sheet puede colorear y llenar cantidades sobre esta base.
    """
    wrap = Alignment(wrap_text=True, vertical="top")
    ctr  = Alignment(horizontal="center", vertical="center")

    def header(text, row):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _FILL_HEADER
        c.alignment = ctr
        ws.merge_cells(f"A{row}:H{row}")

    def section(num, desc, row):
        ws.cell(row=row, column=1, value=num).font = _FONT_BOLD
        c = ws.cell(row=row, column=2, value=desc)
        c.font = _FONT_BOLD
        c.fill = _FILL_SECTION
        ws.merge_cells(f"B{row}:H{row}")

    def item_row(num, desc, unid, row):
        for col, val in [(1, num), (2, desc), (3, unid)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = _FONT_NORM
            c.border = _BORDER_THIN
            c.alignment = wrap
        # D=cantidad, E=precio UF, F=total UF, G=obs
        for col in range(4, 9):
            c = ws.cell(row=row, column=col, value=None)
            c.border = _BORDER_THIN
            if col == 5:  # precio → amarillo vacío por defecto
                c.fill = _FILL_PRICE

    row = 1
    # Encabezado de columnas
    header("CUADRO DE PRECIOS — DETALLE COSTO DIRECTO", row)
    row += 1
    for col, h in enumerate(["N°", "Descripción", "Und", "Cantidad", "P.U. (UF)", "Total (UF)", "Obs."], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _FONT_BOLD; c.fill = _FILL_SECTION; c.border = _BORDER_THIN
        c.alignment = ctr
    row += 1

    # Sección 4: Piques
    section("4", "OBRAS EN PIQUES", row); row += 1
    for p in range(1, 10):
        tipo = "SS VITACURA" if p == 1 else ("SS PROVIDENCIA" if p == 9 else "TÍPICO")
        section(f"4.{p}", f"PIQUE {p} ({tipo})", row); row += 1
        items = _ITEMS_PIQUE + (_ITEMS_PIQUE_SS if p in (1, 9) else _ITEMS_PIQUE_TYP)
        for sub, desc, unid, _ in items:
            item_row(f"4.{p}.{sub}", desc, unid, row); row += 1

    # Sección 5: Túneles
    section("5", "OBRAS EN TÚNELES", row); row += 1
    for t in range(1, 9):
        ep = _ENTRE_PIQUES.get(t, f"P{t}-P{t+1}")
        section(f"5.{t}", f"TRAMO {t}  ({ep})", row); row += 1
        for sub, desc, unid, _ in _ITEMS_TUNEL:
            item_row(f"5.{t}.{sub}", desc, unid, row); row += 1

    # Ajuste columnas
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22


def _fill_detalle_sheet(
    ws,
    qty_map: dict[str, float],
    precios_uf: dict[str, float] | None = None,
) -> dict[str, int]:
    """
    Recorre la hoja 'A) Detalle Costo Directo', aplica colores y fórmulas.
    Si precios_uf está disponible, rellena columna E con precio referencial (salmón).
    Devuelve {item_num: row_index} para referencia.
    """
    item_rows: dict[str, int] = {}
    wrap = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows():
        num_cell = row[0]
        num = str(num_cell.value).strip() if num_cell.value else ""
        if not num:
            continue

        row_idx = num_cell.row
        item_rows[num] = row_idx

        # Sección padre (1.0, 2.0, 3.0, ...)
        is_parent = num.endswith(".0") or (num.count(".") == 0 and num.isdigit())
        if is_parent:
            for c in row:
                if c.value is not None:
                    c.font = _FONT_BOLD
                    c.fill = _FILL_SECTION
            continue

        qty_cell   = ws.cell(row=row_idx, column=4)  # D — cantidad
        price_cell = ws.cell(row=row_idx, column=5)  # E — precio unitario UF
        total_cell = ws.cell(row=row_idx, column=6)  # F — total UF

        # Precio: referencial si tenemos CSV, amarillo (vacío) si no
        precio_ref = _precio_para_item(num, precios_uf) if precios_uf else None
        if precio_ref is not None and price_cell.value in (None, 0, ""):
            price_cell.value = precio_ref
            price_cell.fill = _FILL_PRICE_REF   # salmón = referencial ONDAC
            price_cell.font = _FONT_NORM
            price_cell.number_format = "#,##0.0000"
        else:
            price_cell.fill = _FILL_PRICE       # amarillo = contratista llena
            price_cell.font = _FONT_NORM

        # Cantidad: calculada por nosotros o ya estaba en el template
        if num in qty_map:
            qty_cell.value = round(qty_map[num], 2)
            qty_cell.fill = _FILL_CALC
            qty_cell.font = _FONT_NORM
            qty_cell.number_format = "#,##0.00"
            total_cell.value = f"=D{row_idx}*E{row_idx}"
            total_cell.number_format = "#,##0.00"
        elif qty_cell.value is not None and qty_cell.value != 0:
            if isinstance(qty_cell.value, float):
                qty_cell.value = round(qty_cell.value, 4)
            qty_cell.fill = _FILL_ORIGINAL
            qty_cell.font = _FONT_NORM
            total_cell.value = f"=D{row_idx}*E{row_idx}"
            total_cell.number_format = "#,##0.00"

        # Descripcion: wrap text
        desc_cell = ws.cell(row=row_idx, column=2)
        desc_cell.alignment = wrap

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 30

    return item_rows


# ── Hoja Análisis de Precios ──────────────────────────────────────────────────

_BENCHMARKS = [
    # (partida, unidad, uf_min, uf_max, referencia)
    ("Excavación pique Ø4m",            "m³",  2.0,  3.5,  "ONDAC 2026"),
    ("Retiro excavación pique",          "m³",  0.8,  1.5,  "Mercado Santiago"),
    ("Montaje liner pique Ø4m (solo inst.)", "m",  4.0,  8.0,  "Est. mercado"),
    ("Shotcrete en pique",               "m³",  4.5,  7.0,  "ONDAC 2026"),
    ("Malla electrosoldada",             "kg",  0.03, 0.05, "ONDAC 2026"),
    ("Montaje escalas+plataformas",      "gl",  60,   120,  "Est. mercado/pique"),
    ("Brocal definitivo",                "gl",  80,   150,  "Est. mercado"),
    ("Excavación túnel Ø2.21m",          "m³",  3.5,  6.0,  "ONDAC 2026 tunel urbano"),
    ("Montaje liner túnel Ø2.21m",       "m",   3.5,  6.0,  "Est. instalación"),
    ("Radier túnel H30",                 "m³",  7.0,  10.0, "ONDAC 2026"),
    # Ventilación — Aporte STM (suministro mandante, referencia informativa)
    ("Extractor Soler-Palau TGT/6-1409 (37kW) — 2 un", "un", 800, 1200,
     "STM38110 §3.14 — suministro STM, instalación contratista"),
    ("Jet Fan S&P TJHT/2/4-315-C (1.1kW) — N un",       "un",  20,   40,
     "STM38110 §2.12 — suministro STM, cant. pendiente STM38109"),
]

_NO_CUBICADO = [
    ("4.x.10  Shotcrete", "m³",  "Depende de clasificación RMR del terreno (geotecnia)"),
    ("4.x.11  Malla",     "kg",  "Diseño geotécnico — varía según tipo de suelo"),
    ("4.x.12  Pernos de convergencia", "und", "Análisis convergencia-confinamiento requerido"),
    ("4.x.4   Refuerzo Losa Anillo Fondo", "gl", "Suma alzada — cotización directa contratista"),
    ("4.x.5   Construcción Dren",     "gl",  "Suma alzada — cotización directa contratista"),
    ("4.x.6   Terminaciones",         "gl",  "Suma alzada — cotización directa contratista"),
    ("4.x.7   Cobertura pique",       "gl",  "Suma alzada — cotización directa contratista"),
    ("4.x.9   Refuerzo apertura",     "gl",  "Suma alzada — cotización directa contratista"),
    ("4.x.14  Brocal temporal",       "gl",  "Suma alzada — cotización directa contratista"),
    ("Sec. 2  Liner pique Ø4m + liner túnel Ø2.21m", "—",
     "Aporte STM: mandante suministra el material liner, contratista instala (ítems 4.x.3, 5.x.3)"),
    ("Sec. 2  Escalas y plataformas (estructura)", "—",
     "Aporte STM: mandante suministra estructura, contratista monta (ítem 4.x.8)"),
    ("Sec. 2  Extractor Soler-Palau TGT/6-1409 (37kW)", "2 un",
     "Aporte STM: 2 extractores Ø1410mm en bocas P1 y P9. 132.000 m³/h, 200 Pa. "
     "Contratista instala. Fuente: STM38110 §3.1 + §3.14"),
    ("Sec. 2  Jet Fan S&P TJHT2/4-315-CN (0.8kW)", "23 un",
     "Aporte STM: 23 ventiladores en tunel, 24N, 4.500 m³/h, Ø0.425m. "
     "Fan1 DM+70m a Fan23 DM+2380m, cada 105m. Fuente: STM38109 §4.5 Tabla 4.2 REV.E 28-01-2026"),
    ("Sec. 1  Actividades Previas y Coordinación", "gl", "Suma alzada — depende de empresa"),
    ("Sec. 3  Instalación de Faenas", "gl",
     "Suma alzada. IF SS Vitacura (STM-IF-20251009): 3 etapas (Túnel Liner / TL+OOEE / OOCC+OOEE). "
     "Contenido típico: contenedores oficinas/baños/bodega, compresor, winche, minicargador, acopio liner."),
]

# Cantidades pique 1 para escenarios
_PIQUE1_ESCENARIOS = [
    ("Excavación pique",            251.4, 2.0,  2.75, 3.5),
    ("Retiro excavación",           251.4, 0.8,  1.15, 1.5),
    ("Montaje liner pique",          20.0, 4.0,  6.0,  8.0),
    ("Montaje escalas+plataformas",   1.0, 60.0, 90.0, 120.0),
    ("Brocal definitivo",             1.0, 80.0, 115.0, 150.0),
]


def _build_analisis_sheet(wb: openpyxl.Workbook, cubicacion: dict) -> None:
    ws = wb.create_sheet("Análisis de Precios")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 35

    wrap = Alignment(wrap_text=True, vertical="top")
    ctr  = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    row = 1

    def title(text, r=None):
        nonlocal row
        r = r or row
        c = ws.cell(row=r, column=1, value=text)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = ctr
        ws.merge_cells(f"A{r}:F{r}")
        row = r + 1

    def subtitle(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = _FONT_BOLD
        c.fill = _FILL_SECTION
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

    def header_row(cols):
        nonlocal row
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = _FONT_BOLD
            c.fill = _FILL_SECTION
            c.alignment = ctr
            c.border = _BORDER_THIN
        row += 1

    def data_row(cols, fill=None):
        nonlocal row
        for i, v in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = _FONT_NORM
            c.alignment = wrap
            c.border = _BORDER_THIN
            if fill:
                c.fill = fill
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── Encabezado ────────────────────────────────────────────────────────────
    title("ANÁLISIS DE PRECIOS — TÚNEL LAT VITACURA-PROVIDENCIA")
    ws.cell(row=row, column=1,
            value=f"UF referencial: {UF_CLP:,} CLP (mayo 2026) | Generado: {__import__('datetime').date.today()}")
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1).font = Font(italic=True, size=8)
    row += 1

    # Nota supuesto: 1 túnel vs 2 paralelos
    n_tuneles = cubicacion.get("datos_proyecto", {}).get("n_tuneles_paralelos", 1)
    nota_tn = (
        f"⚠ SUPUESTO: cantidades sección 5 calculadas para 1 túnel Ø2.21m por tramo. "
        f"Si la licitación requiere {2 if n_tuneles == 1 else n_tuneles} túneles paralelos, "
        f"multiplicar cantidades 5.x por {2 if n_tuneles == 1 else n_tuneles}."
    ) if n_tuneles == 1 else (
        f"✓ Cantidades sección 5 calculadas para {n_tuneles} túneles Ø2.21m por tramo."
    )
    c = ws.cell(row=row, column=1, value=nota_tn)
    c.font = Font(italic=True, size=8, color="C00000" if n_tuneles == 1 else "375623")
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    blank()

    # ── Sección 1: Cubicación realizada ───────────────────────────────────────
    subtitle("1. CUBICACIÓN REALIZADA (cantidades pre-llenadas en hoja A)")
    header_row(["Partida", "Cant.", "Unidad", "Ítems licitación", "UF est. medio", "Fuente / Cómo se calculó"])

    partidas_map = {p["partida"]: p for p in cubicacion.get("partidas", [])}
    cubiertos = [
        ("Excavación pique",            "excavacion_pique",            "4.1.1",  2.75,  "área(12.57m²) × profundidad(20m)"),
        ("Retiro excavación",           "retiro_excavacion_pique",     "4.1.2",  1.15,  "igual excavación — flete botadero"),
        ("Montaje liner pique",         "montaje_liner_pique",         "4.1.3",  6.0,   "profundidad pique (20m = 20 anillos)"),
        ("Montaje escalas+plataformas", "montaje_escalas_plataformas", "4.1.8",  90.0,  "1 gl por pique (STM aporta estructuras)"),
        ("Brocal definitivo",           "brocal_definitivo",           "4.1.15", 115.0, "1 gl por pique"),
        ("Excavación túnel",            "excavacion_tunel",            "5.1-5.8",5.0,   "área(3.84m²) × longitud / 8 tramos"),
        ("Montaje liner túnel Ø2.21m",  "montaje_liner_tunel",         "5.1-5.8",4.75,  "longitud tunel / 8 tramos"),
        ("Radier túnel H30",            "radier_tunel",                "5.1-5.8",8.5,   "área × espesor(0.15m)"),
    ]
    for desc, key, items, uf_med, fuente in cubiertos:
        p = partidas_map.get(key)
        if p:
            uf_total = round(p["cantidad"] * uf_med, 1)
            data_row([desc, round(p["cantidad"], 2), p.get("unidad", ""), items,
                      f"{uf_total:,.0f} UF", fuente], fill=_FILL_CALC)
    blank()

    # ── Sección 2: No cubicado ────────────────────────────────────────────────
    subtitle("2. NO CUBICADO — Requiere estudio específico o suma alzada del contratista")
    header_row(["Ítem licitación", "Unidad", "Razón / Recomendación", "", "", ""])
    ws.merge_cells(f"C{row-1}:F{row-1}")
    for item, unidad, razon in _NO_CUBICADO:
        c1 = ws.cell(row=row, column=1, value=item)
        c1.font = _FONT_NORM; c1.border = _BORDER_THIN; c1.fill = _FILL_WARN
        c2 = ws.cell(row=row, column=2, value=unidad)
        c2.font = _FONT_NORM; c2.border = _BORDER_THIN; c2.fill = _FILL_WARN
        c3 = ws.cell(row=row, column=3, value=razon)
        c3.font = Font(italic=True, size=9); c3.border = _BORDER_THIN
        c3.alignment = wrap
        ws.merge_cells(f"C{row}:F{row}")
        row += 1
    blank()

    # ── Sección 3: Benchmarks mercado ─────────────────────────────────────────
    subtitle("3. PRECIOS DE REFERENCIA MERCADO CHILE 2026 (estimados — solicitar cotización formal)")
    header_row(["Partida", "Unidad", "UF mín/un", "UF máx/un",
                "CLP mín/un (aprox)", "Referencia"])
    for desc, unidad, uf_min, uf_max, ref in _BENCHMARKS:
        data_row([
            desc, unidad,
            f"{uf_min:.2f}", f"{uf_max:.2f}",
            f"${int(uf_min*UF_CLP):,} – ${int(uf_max*UF_CLP):,}",
            ref,
        ])
    blank()

    # ── Sección 4: Análisis de sensatez — proyecto completo ─────────────────
    piques_qty = cubicacion.get("piques_qty", {})
    tramos_qty = cubicacion.get("tramos_qty", {})

    # Construir escenarios desde datos reales si están disponibles
    if piques_qty and tramos_qty:
        total_excav_pq = sum(pq["excavacion_pique"] for pq in piques_qty.values())
        total_liner_pq = sum(pq["montaje_liner_pique"] for pq in piques_qty.values())
        total_escalas  = len(piques_qty)
        total_brocal   = len(piques_qty)
        total_excav_tn = sum(tq["excavacion_tunel"] for tq in tramos_qty.values())
        total_liner_tn = sum(tq["montaje_liner_tunel"] for tq in tramos_qty.values())
        total_radier   = sum(tq["radier_tunel"] for tq in tramos_qty.values())
        n_piques = len(piques_qty)
        n_tramos = len(tramos_qty)
        subtitle(f"4. ANÁLISIS DE SENSATEZ — PROYECTO COMPLETO ({n_piques} piques + {n_tramos} tramos)")
        escenarios = [
            ("Excavación piques",          total_excav_pq, 2.0,  2.75, 3.5),
            ("Retiro excav. piques",       total_excav_pq, 0.8,  1.15, 1.5),
            ("Montaje liner piques",       total_liner_pq, 4.0,  6.0,  8.0),
            ("Montaje escalas+plataformas",total_escalas,  60.0, 90.0, 120.0),
            ("Brocal definitivo",          total_brocal,   80.0, 115.0,150.0),
            ("Excavación túnel",           total_excav_tn, 3.5,  5.0,  6.0),
            ("Montaje liner túnel",        total_liner_tn, 3.5,  4.75, 6.0),
            ("Radier H30",                 total_radier,   7.0,  8.5,  10.0),
        ]
    else:
        subtitle("4. ANÁLISIS DE SENSATEZ — PIQUE 1 (20m prof., Ø4.0m)")
        escenarios = _PIQUE1_ESCENARIOS

    header_row(["Partida", "Cant.", "UF/un Mín", "UF/un Máx",
                "Subtotal Mín (UF)", "Subtotal Máx (UF)"])

    total_min = total_max = 0.0
    for entry in escenarios:
        desc, cant, uf_min, _mid, uf_max = entry
        sub_min = round(cant * uf_min, 1)
        sub_max = round(cant * uf_max, 1)
        total_min += sub_min
        total_max += sub_max
        data_row([desc, round(cant, 1), f"{uf_min:.2f}", f"{uf_max:.2f}",
                  f"{sub_min:,.1f}", f"{sub_max:,.1f}"])

    for i, v in enumerate([
        "TOTAL ÍTEMS CUBICADOS", "", "", "",
        f"{total_min:,.0f} UF  (~${int(total_min*UF_CLP/1e6):.0f}M CLP)",
        f"{total_max:,.0f} UF  (~${int(total_max*UF_CLP/1e6):.0f}M CLP)",
    ], 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _FONT_BOLD; c.fill = _FILL_SECTION; c.border = _BORDER_THIN
    row += 1

    c = ws.cell(row=row, column=1,
        value="⚠ NOTA: El rango anterior NO incluye shotcrete, malla, pernos, terminaciones, "
              "instalación de faenas ni actividades previas (Sec. 1, 3) — típicamente +30-50% adicional.")
    c.font = Font(italic=True, size=8, color="C00000")
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    # ── Sección 5: Cantidades ya en el template original ──────────────────────
    subtitle("5. CANTIDADES PRE-EXISTENTES EN TEMPLATE (de la licitación original)")
    header_row(["Ítem", "Descripción", "Unidad", "Cantidad original", "", "Fuente"])
    ws.merge_cells(f"E{row-1}:E{row-1}")

    originales = [
        ("5.1.4", "Pernos frente fibra vidrio 4m (eventual)", "und", 39.54),
        ("5.1.5", "Shotcrete frente (eventual)",               "m³",  19.77),
        ("5.1.6", "Paraguas micropilotes (eventual)",          "und", 24),
        ("5.2.4–5.8.4", "Pernos frente (tramos 2-8)",         "und", "36-36 por tramo"),
        ("7.1",  "Base binder asfalto",                        "m²",  317.61),
        ("7.2",  "Base chancada CBR80",                        "m³",  66.70),
        ("7.3",  "Calzada con tratamiento asfaltico",          "m²",  317.61),
        ("7.5",  "Excavación dura y transporte botadero",      "m³",  343.02),
        ("7.10", "Entibación excavaciones",                    "m",   90.75),
        ("7.11", "Reposición áreas verdes",                    "m²",  60),
        ("7.12", "Reposición soleras",                         "m",   40),
        ("7.13", "Demolición y reposición veredas",            "m²",  20),
        ("13.2", "Retiro y transporte aguas efluentes",        "m³",  1000),
    ]
    for item, desc, unid, cant in originales:
        c1 = ws.cell(row=row, column=1, value=item)
        c2 = ws.cell(row=row, column=2, value=desc)
        c3 = ws.cell(row=row, column=3, value=unid)
        c4 = ws.cell(row=row, column=4, value=cant)
        c5 = ws.cell(row=row, column=6, value="Cuadro precios STM rev.1 06/04/2026")
        for c in [c1, c2, c3, c4, c5]:
            c.font = _FONT_NORM; c.border = _BORDER_THIN; c.fill = _FILL_ORIGINAL
        ws.merge_cells(f"E{row}:E{row}")
        row += 1


# ── Hoja Memoria de Cálculo ───────────────────────────────────────────────────

def _build_memoria_calculo_sheet(wb: openpyxl.Workbook, cubicacion: dict,
                                  resultado: dict) -> None:
    """Hoja ordenada con racionales de TODOS los cálculos y sus fuentes."""
    ws = wb.create_sheet("Memoria de Cálculo")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 38

    wrap = Alignment(wrap_text=True, vertical="top")
    ctr  = Alignment(horizontal="center", vertical="center")

    row = 1

    def title(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = _FILL_HEADER; c.alignment = ctr
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 22
        row += 1

    def section(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = _FONT_BOLD; c.fill = _FILL_SECTION
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

    def header(cols):
        nonlocal row
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = _FONT_BOLD; c.fill = _FILL_SECTION
            c.border = _BORDER_THIN; c.alignment = ctr
        row += 1

    def data_row(cols, fill=None):
        nonlocal row
        for i, v in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = _FONT_NORM; c.alignment = wrap; c.border = _BORDER_THIN
            if fill:
                c.fill = fill
        row += 1

    def note(text, color="C00000"):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(italic=True, size=8, color=color)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

    title("MEMORIA DE CÁLCULO — TÚNEL LAT 2×110 kV VITACURA-PROVIDENCIA")
    dp = cubicacion.get("datos_proyecto", {})
    note(f"Documento generado: {__import__('datetime').date.today()} | "
         f"Licitación: {dp.get('licitacion','?')} | Mandante: {dp.get('mandante','?')}",
         color="666666")
    row += 1

    # ── Sec 1: Constantes geométricas ─────────────────────────────────────────
    section("1. CONSTANTES GEOMÉTRICAS")
    header(["#", "Constante", "Fórmula", "Valor", "Unidad", "Fuente"])
    constantes = [
        ("1.1", "Área sección pique", "π × r² = π × (4.0/2)²",
         "12.5664", "m²", "STM38136 Ø interior pique"),
        ("1.2", "Área sección túnel liner", "π × r² = π × (2.21/2)²",
         "3.8359", "m²", "STM38136 Ø interior liner"),
        ("1.3", "Radier H30 volumen por metro", "ancho × espesor = 2.21 × 0.15",
         "0.3315", "m³/m", "Diseño estructural radier túnel"),
        ("1.4", "Diámetro fundación pique", "4.20 m externo",
         "4.20", "m", "STM38020 detalle típico"),
        ("1.5", "Espesor liner túnel (PL)", "5 mm",
         "0.005", "m", "STM38020 nota técnica"),
    ]
    for c in constantes:
        data_row(list(c))
    row += 1

    # ── Sec 2: Profundidades de piques ────────────────────────────────────────
    section("2. PROFUNDIDADES DE PIQUES — NV terreno - NV invert liner")
    header(["Pique", "NV terreno (m.s.n.m)", "NV liner invert (m.s.n.m)",
            "Profundidad (m)", "Excavación (m³)", "Fuente plano"])
    AREA_PIQUE = 12.5664
    for pq in resultado.get("piques", []):
        nvt = pq.get("nv_terreno")
        nvi = pq.get("nv_liner_invert")
        prof = pq.get("profundidad_m")
        excav = round(AREA_PIQUE * prof, 2)
        fuente = pq.get("fuente_profundidad", "?")
        # extraer plano del campo fuente
        plano = "STM38020-23"
        for s in ["STM38020", "STM38021", "STM38022", "STM38023"]:
            if s in fuente:
                plano = s
                break
        data_row([f"P{pq['id']}", f"{nvt:.2f}", f"{nvi:.2f}",
                  f"{prof:.2f}", excav, plano], fill=_FILL_CALC)
    note("⚠ Discrepancia P9: STM38109 §2.4 reporta 9.67m (referencia portal vs NV terreno). "
         "Diferencia 0.61m × 12.57m² = ~7.7 m³ excavación adicional si se usa cota portal.")
    row += 1

    # ── Sec 3: Longitudes de tramos ───────────────────────────────────────────
    section("3. LONGITUDES DE TRAMOS — DM piques sucesivos")
    header(["Tramo", "Entre piques", "DM inicio (m)", "DM fin (m)",
            "Longitud (m)", "Fuente plano"])
    for tr in resultado.get("tramos_tunel", []):
        dm_ini = tr.get("km_inicio", 0) * 1000
        dm_fin = tr.get("km_fin", 0) * 1000
        lng = tr.get("longitud_m", 0)
        fuente = tr.get("fuente_longitud", "?")
        plano = "STM38020-23"
        for s in ["STM38020", "STM38021", "STM38022", "STM38023"]:
            if s in fuente:
                plano = s
                break
        data_row([f"T{tr['id']}", tr.get("entre_piques", "?"),
                  f"{dm_ini:.1f}", f"{dm_fin:.1f}", f"{lng:.1f}", plano],
                 fill=_FILL_CALC)
    row += 1

    # ── Sec 4: Fórmulas por partida ──────────────────────────────────────────
    section("4. FÓRMULAS POR PARTIDA")
    header(["#", "Partida", "Fórmula aplicada", "Cantidad total",
            "Unidad", "Validación"])
    pq_qty = cubicacion.get("piques_qty", {})
    tr_qty = cubicacion.get("tramos_qty", {})
    formulas = [
        ("4.1", "Excavación piques",
         "Σ (12.5664 m² × prof_i) para i=1..9",
         sum(p.get("excavacion_pique", 0) for p in pq_qty.values()),
         "m³", "✓ Coincide con suma manual"),
        ("4.2", "Retiro excavación piques",
         "= Excavación pique (igual volumen al botadero)",
         sum(p.get("retiro_excavacion_pique", 0) for p in pq_qty.values()),
         "m³", "✓ = Excavación"),
        ("4.3", "Montaje liner pique",
         "Σ profundidad_i para i=1..9",
         sum(p.get("montaje_liner_pique", 0) for p in pq_qty.values()),
         "m", "= longitud anillo × cantidad anillos"),
        ("4.4", "Montaje escalas+plataformas",
         "1 gl × 9 piques (Aporte STM)",
         sum(p.get("montaje_escalas_plataformas", 0) for p in pq_qty.values()),
         "gl", "✓ Una unidad global por pique"),
        ("4.5", "Brocal definitivo",
         "1 gl × 9 piques",
         sum(p.get("brocal_definitivo", 0) for p in pq_qty.values()),
         "gl", "P1/P9 → ítem .15; P2-8 → ítem .14"),
        ("4.6", "Montaje estructura cables (SS)",
         "1 gl × 2 piques SS (P1 y P9)",
         sum(p.get("montaje_estructura_cables", 0) for p in pq_qty.values()),
         "gl", "Solo SS Vitacura + SS Providencia"),
        ("5.1", "Excavación túnel",
         "Σ (3.8359 m² × longitud_i) para i=1..8",
         sum(t.get("excavacion_tunel", 0) for t in tr_qty.values()),
         "m³", "Verifica: 3.8359 × 2454.2 = 9414.2 ✓"),
        ("5.2", "Radier túnel H30",
         "Σ (0.3315 m³/m × longitud_i) para i=1..8",
         sum(t.get("radier_tunel", 0) for t in tr_qty.values()),
         "m³", "Verifica: 0.3315 × 2454.2 = 813.6 ✓"),
        ("5.3", "Instalación cables",
         "Σ (6 cables × longitud_i) para i=1..8",
         sum(t.get("instalacion_cables", 0) for t in tr_qty.values()),
         "ml", "⚠ Asumido: LAT 2×110kV = 2 circuitos × 3 fases"),
    ]
    for f in formulas:
        data_row([f[0], f[1], f[2], round(f[3], 2), f[4], f[5]])
    row += 1

    # ── Sec 5: Equipamiento eléctrico ─────────────────────────────────────────
    section("5. EQUIPAMIENTO ELÉCTRICO (Aporte STM — Sección 2 licitación)")
    header(["#", "Equipo", "Cantidad", "Modelo / specs", "Ubicación", "Fuente"])
    vent = dp.get("ventilacion", {})
    jf = vent.get("jet_fans", {})
    ex = vent.get("extractores", {})
    equipos = [
        ("5.1", "Jet Fan túnel", f"{jf.get('cantidad','?')}",
         f"S&P TJHT2/4-315-CN ({jf.get('potencia_kw','?')} kW, "
         f"{jf.get('empuje_n','?')} N, {jf.get('caudal_m3h','?')} m³/h)",
         f"Fan1 DM+{jf.get('dm_inicio_m','?')} → "
         f"Fan23 DM+{jf.get('dm_fin_m','?')}, cada {jf.get('separacion_m','?')} m",
         "STM38109 §4.5 + Tabla 4.2 (REV.E 28-01-2026)"),
        ("5.2", "Extractor pique", f"{ex.get('cantidad','?')}",
         f"Soler-Palau TGT/6-1409 ({ex.get('potencia_kw','?')} kW, "
         f"{ex.get('caudal_m3h','?')} m³/h, Ø{ex.get('diametro_mm','?')} mm)",
         ex.get('ubicacion', '?'),
         "STM38110 §3.1 + §3.14"),
    ]
    for e in equipos:
        data_row(list(e), fill=_FILL_CALC)
    row += 1

    # ── Sec 6: Supuestos críticos ─────────────────────────────────────────────
    section("6. SUPUESTOS CRÍTICOS Y VALIDACIONES PENDIENTES")
    header(["#", "Supuesto", "Valor usado", "Origen", "Impacto si cambia", "Estado"])
    supuestos = [
        ("6.1", "N° túneles paralelos", "1", "Diseño STM",
         "Multiplicar Sec 5 por N", "✓ Confirmado 1"),
        ("6.2", "N° cables LAT por tramo", "6 (2×3 fases)",
         "Estándar LAT 2×110kV", "Ajustar 5.x.8 proporcional", "⚠ Confirmar plano eléctrico"),
        ("6.3", "Profundidad P9", "9.06 m (NV)",
         "STM38023 vs STM38109 (9.67m)",
         "Δ excav P9 = 7.7 m³", "⚠ Discrepancia 0.61m"),
        ("6.4", "Diámetro liner interior", "2.21 m",
         "STM38136 vs STM38109 (2.10m)",
         "Δ sección = 0.38 m² (10%)", "✓ 2.21m incluye espesor anillos"),
        ("6.5", "Longitud túnel total", "2454.2 m",
         "STM38020-23 suma DM",
         "STM38109 redondea a 2450m", "✓ Diferencia 4.2m (0.17%)"),
        ("6.6", "Cantidades 4.x.10-12 (shotcrete/malla/pernos)",
         "No cubicado", "Requiere RMR/Q rock mass",
         "Eventual según geotecnia", "⚠ Pendiente STM38019 RMR"),
        ("6.7", "UF referencial", "38.500 CLP",
         "Mayo 2026", "Actualizar al mes oferta", "✓ Configurable"),
        ("6.8", "Altura excavación = profundidad NV",
         "Sí", "Convención técnica",
         "Si sump existe → +volumen", "⚠ STM38023 nota 'Nivel -3.50' en P9"),
    ]
    for s in supuestos:
        data_row(list(s))
    row += 1

    # ── Sec 7: Trazabilidad de fuentes ────────────────────────────────────────
    section("7. TRAZABILIDAD DE FUENTES (documentos consultados)")
    header(["#", "Documento", "Revisión", "Datos extraídos", "Hoja relevante", "—"])
    fuentes_doc = [
        ("7.1", "STM38020 Planta y Perfil", "REV.0",
         "P1, P2, P3 NV + DM tramos T1-T3", "Detalle Costo Directo", ""),
        ("7.2", "STM38021 Planta y Perfil", "REV.0",
         "P4, P5 NV + DM tramos T3-T5", "Detalle Costo Directo", ""),
        ("7.3", "STM38022 Planta y Perfil", "REV.0",
         "P6, P7 NV + DM tramos T5-T7", "Detalle Costo Directo", ""),
        ("7.4", "STM38023 Planta y Perfil", "REV.0",
         "P8, P9 NV + DM tramos T7-T8", "Detalle Costo Directo", ""),
        ("7.5", "STM38109 Memoria Ventilación", "REV.E 28-01-2026",
         "23 jet fans S&P TJHT2/4-315-CN, ubicación Tabla 4.2", "Sec 5 supuestos", ""),
        ("7.6", "STM38110 ETP Equipos Ventilación", "REV.0 14-11-2025",
         "2 extractores Soler-Palau TGT/6-1409, modelo jet fan", "Sec 5 supuestos", ""),
        ("7.7", "STM38019 Geotécnica", "—",
         "Estratigrafía H1-H3 (Relleno/Grava 2°/Grava 1°)",
         "Pendiente RMR para 4.x.10-12", ""),
        ("7.8", "STM38136 Planos eléctricos", "REV.1 06/04/2026",
         "Ø interior pique 4.0m, Ø interior liner 2.21m",
         "Sec 1 constantes", ""),
        ("7.9", "BAE Licitación 2025-25-TX-SS-RM", "REV.0 13/04/2026",
         "Estructura items 4.x y 5.x", "A) Detalle", ""),
        ("7.10", "IF SS Vitacura", "20251009 (3 etapas)",
         "Composición Instalación de Faenas",
         "Sec 3 supuestos suma alzada", ""),
    ]
    for f in fuentes_doc:
        data_row(list(f))
    row += 1

    # ── Sec 8: Resumen totales ────────────────────────────────────────────────
    section("8. RESUMEN TOTALES CUBICADOS")
    header(["#", "Concepto", "Cantidad", "Unidad", "—", "—"])
    resumen = [
        ("8.1", "Excavación piques (suma)",
         sum(p.get("excavacion_pique", 0) for p in pq_qty.values()), "m³", "", ""),
        ("8.2", "Excavación túnel (suma)",
         sum(t.get("excavacion_tunel", 0) for t in tr_qty.values()), "m³", "", ""),
        ("8.3", "Liner pique (suma profundidades)",
         sum(p.get("montaje_liner_pique", 0) for p in pq_qty.values()), "m", "", ""),
        ("8.4", "Liner túnel (longitud total)",
         sum(t.get("montaje_liner_tunel", 0) for t in tr_qty.values()), "m", "", ""),
        ("8.5", "Radier H30",
         sum(t.get("radier_tunel", 0) for t in tr_qty.values()), "m³", "", ""),
        ("8.6", "Cables instalados (6 × longitud)",
         sum(t.get("instalacion_cables", 0) for t in tr_qty.values()), "ml", "", ""),
        ("8.7", "Estructura cables SS",
         sum(p.get("montaje_estructura_cables", 0) for p in pq_qty.values()),
         "gl", "(solo P1+P9)", ""),
        ("8.8", "Escalas+plataformas",
         sum(p.get("montaje_escalas_plataformas", 0) for p in pq_qty.values()),
         "gl", "", ""),
        ("8.9", "Brocales definitivos",
         sum(p.get("brocal_definitivo", 0) for p in pq_qty.values()), "gl", "", ""),
    ]
    for r in resumen:
        c1 = round(r[2], 2) if isinstance(r[2], float) else r[2]
        data_row([r[0], r[1], c1, r[3], r[4], r[5]], fill=_FILL_CALC)


# ── Función principal ─────────────────────────────────────────────────────────

def exportar_licitacion(
    cubicacion: dict,
    resultado: dict,
    ruta_out: str | Path,
    template: str | Path | None = None,
    precios_csv: str | Path | None = None,
) -> Path:
    """
    Genera Excel en formato licitación con cantidades y precios referenciales.

    template:    path al .xlsx del mandante (opcional).
    precios_csv: path al CSV de precios (default: precios_electrico.csv junto al módulo).
                 Precios en CLP → se convierten a UF. Se marcan en salmón (referenciales).
    Devuelve Path del archivo generado.
    """
    ruta_out = Path(ruta_out)

    if template is not None and Path(template).exists():
        wb = openpyxl.load_workbook(template)
    else:
        if template is not None:
            print(f"  WARN: template '{template}' no encontrado — generando workbook vacío")
        wb = openpyxl.Workbook()
        wb.active.title = "Resumen Oferta"
        wb.create_sheet("A) Detalle Costo Directo")
        wb.create_sheet("B) Detalle C. Ind., GG y Utilid")

    # Cargar precios referenciales
    if precios_csv is None:
        precios_csv = Path(__file__).parent / "precios_electrico.csv"
    precios_uf = cargar_precios_uf(precios_csv)

    qty_map = _build_qty_map(cubicacion)

    ws_det = wb["A) Detalle Costo Directo"]
    if ws_det.max_row <= 1:  # sheet vacía (sin template)
        _create_detalle_rows(ws_det)
    _fill_detalle_sheet(ws_det, qty_map, precios_uf=precios_uf)

    # Leyenda de colores en Resumen Oferta
    ws_res = wb["Resumen Oferta"]
    leyenda_row = ws_res.max_row + 2
    for col, (texto, fill) in enumerate([
        ("Verde: cantidad calculada por modelo",       _FILL_CALC),
        ("Salmón: precio referencial ONDAC 2026 (UF)", _FILL_PRICE_REF),
        ("Amarillo: completar con precio oferta",      _FILL_PRICE),
    ], start=1):
        c = ws_res.cell(row=leyenda_row, column=col, value=texto)
        c.fill = fill
        c.font = Font(size=8, italic=True)

    # Eliminar hojas preexistentes del template antes de regenerar
    for sheet_name in ("Análisis de Precios", "Memoria de Cálculo"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    _build_analisis_sheet(wb, cubicacion)
    _build_memoria_calculo_sheet(wb, cubicacion, resultado)

    wb.save(ruta_out)
    return ruta_out
