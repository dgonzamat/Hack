"""
Tests unitarios e integrados para cubicador.py, presupuesto.py y excel.py.

Datos sintéticos DEPTO B1:
  COCINA-LAV  7.6 m²  húmedo
  DORM.2      9.0 m²  seco
  DORM.PRIN  11.2 m²  seco
  LIVING     22.8 m²  seco
  LOGIA       4.5 m²  húmedo
  WALK-IN     3.2 m²  seco
"""

import math
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cubicador import (
    clasificar_recinto,
    perimetro_recinto,
    area_efectiva,
    dim_vano,
    area_vanos_total,
    cubicar,
    cubicar_vial,
)
from presupuesto import cargar_precios, presupuestar
from excel import exportar_excel


# ─── Fixtures ─────────────────────────────────────────────────────────────────

RECINTOS_B1 = [
    {"nombre": "COCINA-LAV",  "area_m2": 7.6,  "confianza": 0.9, "departamento": "B1"},
    {"nombre": "DORM.2",      "area_m2": 9.0,  "confianza": 0.9, "departamento": "B1"},
    {"nombre": "DORM.PRIN",   "area_m2": 11.2, "confianza": 0.9, "departamento": "B1"},
    {"nombre": "LIVING",      "area_m2": 22.8, "confianza": 0.9, "departamento": "B1"},
    {"nombre": "LOGIA",       "area_m2": 4.5,  "confianza": 0.9, "departamento": "B1"},
    {"nombre": "WALK-IN",     "area_m2": 3.2,  "confianza": 0.9, "departamento": "B1"},
]

RESULTADO_B1 = {
    "recintos": RECINTOS_B1,
    "muros": {"interior_ml": 20.0, "exterior_ml": 18.0},
    "vanos": {
        "puertas": [{"tipo": "puerta simple 90", "cantidad": 5}],
        "ventanas": [{"tipo": "ventana corredera 120", "cantidad": 2}],
    },
    "lamina": {"titulo": "Planta B1 Sintetica", "tipo": "planta", "escala": "1:50"},
    "calidad_plano": "alta",
    "observaciones": "",
    "_tiles_procesados": 1,
    "_tokens_entrada": 1000,
    "_tokens_salida": 500,
}

PRECIOS_CSV = Path(__file__).parent / "precios_cl.csv"


# ─── clasificar_recinto ────────────────────────────────────────────────────────

class TestClasificarRecinto:
    def test_humedo_cocina(self):
        cat, heur = clasificar_recinto("COCINA")
        assert cat == "humedo" and not heur

    def test_humedo_bano(self):
        cat, _ = clasificar_recinto("BAÑO PRINCIPAL")
        assert cat == "humedo"

    def test_humedo_logia(self):
        cat, _ = clasificar_recinto("LOGIA")
        assert cat == "humedo"

    def test_humedo_compuesto(self):
        # COCINA-LAVANDERIA debe clasificar como húmedo
        cat, _ = clasificar_recinto("COCINA-LAV")
        assert cat == "humedo"

    def test_seco_dormitorio(self):
        cat, heur = clasificar_recinto("DORM.PRINCIPAL")
        assert cat == "seco" and not heur

    def test_seco_living(self):
        cat, _ = clasificar_recinto("LIVING")
        assert cat == "seco"

    def test_seco_walk_in(self):
        cat, _ = clasificar_recinto("WALK-IN CLOSET")
        assert cat == "seco"

    def test_exterior_terraza(self):
        cat, _ = clasificar_recinto("TERRAZA")
        assert cat == "exterior"

    def test_comun_hall(self):
        cat, _ = clasificar_recinto("HALL DE DISTRIBUCION")
        assert cat == "comun"

    def test_comun_escalera(self):
        cat, _ = clasificar_recinto("ESCALERA")
        assert cat == "comun"

    def test_bodega_es_seco(self):
        cat, _ = clasificar_recinto("BODEGA")
        assert cat == "seco"

    def test_nombre_desconocido_heuristico(self):
        # Un nombre completamente desconocido debe tener confianza baja → fue_heuristica
        cat, heur = clasificar_recinto("ZAGUÁN XKQZ")
        assert cat in ("seco", "comun", "humedo", "exterior", "vial")
        assert heur  # confianza ML < 0.40 para nombre absurdo

    def test_vial_calzada(self):
        cat, _ = clasificar_recinto("CALZADA")
        assert cat == "vial"

    def test_vial_tunel(self):
        cat, _ = clasificar_recinto("TUNEL")
        assert cat == "vial"

    def test_vial_vereda(self):
        cat, _ = clasificar_recinto("VEREDA")
        assert cat == "vial"

    def test_vial_portal(self):
        cat, _ = clasificar_recinto("PORTAL NORTE")
        assert cat == "vial"

    def test_vial_cuneta(self):
        cat, _ = clasificar_recinto("CUNETA HORMIGON")
        assert cat == "vial"

    def test_vial_puente(self):
        cat, _ = clasificar_recinto("PUENTE VEHICULAR")
        assert cat == "vial"

    def test_vial_metro(self):
        cat, _ = clasificar_recinto("ANDEN METRO")
        assert cat == "vial"

    def test_word_boundary_no_falso_positivo(self):
        # "TALLER" no debe matchear "HALL" (sin word boundary sería un bug)
        cat, _ = clasificar_recinto("TALLER")
        assert cat != "comun"


# ─── perimetro_recinto ─────────────────────────────────────────────────────────

class TestPerimetroRecinto:
    def test_con_dimensiones_exactas(self):
        rec = {"area_m2": 20.0, "dimensiones_estimadas": {"largo_m": 5.0, "ancho_m": 4.0}}
        perim, estimado = perimetro_recinto(rec)
        assert perim == pytest.approx(18.0, abs=0.01)
        assert not estimado

    def test_fallback_rectangulo_15_1(self):
        # Para A=12 m²: L=sqrt(18)=4.243, W=sqrt(8)=2.828 → P=2*(4.243+2.828)=14.14
        rec = {"area_m2": 12.0}
        perim, estimado = perimetro_recinto(rec)
        L = math.sqrt(1.5 * 12)
        W = math.sqrt(12 / 1.5)
        esperado = 2 * (L + W)
        assert perim == pytest.approx(esperado, abs=0.01)
        assert estimado

    def test_area_none_devuelve_cero(self):
        rec = {"area_m2": None}
        perim, estimado = perimetro_recinto(rec)
        assert perim == 0.0 and estimado

    def test_area_cero_devuelve_cero(self):
        rec = {"area_m2": 0}
        perim, estimado = perimetro_recinto(rec)
        assert perim == 0.0

    def test_fallback_no_subestima_cuadrado(self):
        # Rectángulo 1.5:1 debe dar perímetro mayor que cuadrado (P_cuadrado = 4*sqrt(A))
        # pero la diferencia debe ser < 2% para A razonable
        area = 12.0
        rec = {"area_m2": area}
        perim, _ = perimetro_recinto(rec)
        p_cuadrado = 4 * math.sqrt(area)
        # Rectángulo 1.5:1 ≈ 2% más que cuadrado — no subestima
        assert perim >= p_cuadrado * 0.99

    @pytest.mark.parametrize("area,esperado_aprox", [
        (7.6,  11.25),
        (9.0,  12.25),
        (11.2, 13.66),
        (22.8, 19.49),
        (4.5,   8.66),
        (3.2,   7.30),
    ])
    def test_perimetros_b1(self, area, esperado_aprox):
        rec = {"area_m2": area}
        perim, _ = perimetro_recinto(rec)
        assert perim == pytest.approx(esperado_aprox, abs=0.05)


# ─── area_efectiva ─────────────────────────────────────────────────────────────

class TestAreaEfectiva:
    def test_usa_area_m2(self):
        assert area_efectiva({"area_m2": 15.5}) == pytest.approx(15.5)

    def test_fallback_dimensiones(self):
        rec = {"area_m2": None, "dimensiones_estimadas": {"largo_m": 4.0, "ancho_m": 3.0}}
        assert area_efectiva(rec) == pytest.approx(12.0)

    def test_ninguno_devuelve_none(self):
        assert area_efectiva({"nombre": "X"}) is None


# ─── dim_vano ─────────────────────────────────────────────────────────────────

class TestDimVano:
    def test_puerta_simple(self):
        a, h, default = dim_vano("puerta simple 90")
        assert a == pytest.approx(0.90) and h == pytest.approx(2.00) and not default

    def test_puerta_doble(self):
        a, h, default = dim_vano("puerta doble 150")
        assert a == pytest.approx(1.50) and not default

    def test_ventana_corredera(self):
        a, h, default = dim_vano("ventana corredera 120")
        assert a == pytest.approx(1.20) and h == pytest.approx(1.00) and not default

    def test_ventana_fija(self):
        a, h, default = dim_vano("ventana fija")
        assert a == pytest.approx(1.50) and not default

    def test_tipo_desconocido_devuelve_default(self):
        a, h, default = dim_vano("acceso peatonal")
        assert default

    def test_tipo_vacio(self):
        a, h, default = dim_vano("")
        assert default


# ─── cubicar (integrado B1) ───────────────────────────────────────────────────

class TestCubicarB1:
    @pytest.fixture(scope="class")
    def cubicacion(self):
        return cubicar(RESULTADO_B1, altura_global=2.4)

    def test_pintura_interior_rango(self, cubicacion):
        """Pintura interior esperada ~325.8 m² (±5% = 309-342 m²)"""
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        q = partidas["pintura_interior_latex_2manos"]
        assert 309 <= q <= 345, f"Pintura interior {q} m² fuera del rango [309, 345]"

    def test_piso_humedo(self, cubicacion):
        """COCINA-LAV (7.6) + LOGIA (4.5) = 12.1 m²"""
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        assert partidas["piso_ceramico_60x60_instalado"] == pytest.approx(12.1, abs=0.1)

    def test_piso_seco(self, cubicacion):
        """DORM.2 + DORM.PRIN + LIVING + WALK-IN = 9.0+11.2+22.8+3.2 = 46.2 m²"""
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        assert partidas["piso_flotante_8mm_instalado"] == pytest.approx(46.2, abs=0.1)

    def test_cielo(self, cubicacion):
        """Todos los 6 recintos = 58.3 m²"""
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        assert partidas["cielo_volcanita_pintado"] == pytest.approx(58.3, abs=0.1)

    def test_puertas_conteo(self, cubicacion):
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        assert partidas["puerta_simple_90cm_instalada"] == 5

    def test_puerta_doble_separada(self):
        res = dict(RESULTADO_B1)
        res["vanos"] = {
            "puertas": [
                {"tipo": "puerta simple 90", "cantidad": 3},
                {"tipo": "puerta doble 150", "cantidad": 2},
            ],
            "ventanas": [],
        }
        cub = cubicar(res, altura_global=2.4)
        partidas = {p["partida"]: p["cantidad"] for p in cub["partidas"]}
        assert partidas.get("puerta_simple_90cm_instalada") == 3
        assert partidas.get("puerta_doble_150cm_instalada") == 2

    def test_ventanas_area(self, cubicacion):
        """2 ventanas corredera 120 → 2 × 1.20 × 1.00 = 2.4 m²"""
        partidas = {p["partida"]: p["cantidad"] for p in cubicacion["partidas"]}
        assert partidas["ventana_corredera_aluminio"] == pytest.approx(2.4, abs=0.1)

    def test_clasificacion_correcta(self, cubicacion):
        by_nombre = {c["nombre"]: c["categoria"] for c in cubicacion["clasificacion_recintos"]}
        assert by_nombre["COCINA-LAV"] == "humedo"
        assert by_nombre["LOGIA"] == "humedo"
        assert by_nombre["DORM.2"] == "seco"
        assert by_nombre["DORM.PRIN"] == "seco"
        assert by_nombre["LIVING"] == "seco"
        assert by_nombre["WALK-IN"] == "seco"

    def test_sin_recintos_sin_area(self, cubicacion):
        assert cubicacion["recintos_sin_area"] == []

    def test_resumen_areas(self, cubicacion):
        r = cubicacion["resumen_areas"]
        assert r["total_m2"] == pytest.approx(58.3, abs=0.1)
        assert r["incierta_m2"] == 0.0  # todos conf=0.9

    def test_terraza_no_entra_a_cubicacion(self):
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "TERRAZA", "area_m2": 10.0, "confianza": 0.9, "departamento": "B1"}
        ]
        cub = cubicar(res, altura_global=2.4)
        partidas = {p["partida"]: p["cantidad"] for p in cub["partidas"]}
        # Cielo no incluye terraza
        assert partidas["cielo_volcanita_pintado"] == pytest.approx(58.3, abs=0.1)

    def test_comun_excluido_por_default(self):
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "HALL", "area_m2": 5.0, "confianza": 0.9, "departamento": "B1"}
        ]
        cub = cubicar(res, altura_global=2.4)
        assert len(cub["comunes_omitidos"]) == 1
        partidas = {p["partida"]: p["cantidad"] for p in cub["partidas"]}
        assert partidas["cielo_volcanita_pintado"] == pytest.approx(58.3, abs=0.1)

    def test_comun_incluido_con_flag(self):
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "HALL", "area_m2": 5.0, "confianza": 0.9, "departamento": "B1"}
        ]
        cub = cubicar(res, altura_global=2.4, incluir_comunes=True)
        assert cub["comunes_omitidos"] == []
        partidas = {p["partida"]: p["cantidad"] for p in cub["partidas"]}
        assert partidas["cielo_volcanita_pintado"] == pytest.approx(63.3, abs=0.1)

    def test_recinto_sin_area_omitido(self):
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "BODEGA", "area_m2": None, "confianza": 0.8, "departamento": "B1"}
        ]
        cub = cubicar(res, altura_global=2.4)
        assert "BODEGA" in cub["recintos_sin_area"]

    def test_vial_excluido_de_cubicacion(self):
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "CALZADA", "area_m2": 500.0, "confianza": 0.9, "departamento": None},
            {"nombre": "TUNEL", "area_m2": 200.0, "confianza": 0.9, "departamento": None},
        ]
        cub = cubicar(res, altura_global=2.4)
        assert len(cub["viales_detectados"]) == 2
        # Cielo no debe incluir calzada ni tunel
        partidas = {p["partida"]: p["cantidad"] for p in cub["partidas"]}
        assert partidas["cielo_volcanita_pintado"] == pytest.approx(58.3, abs=0.1)

    def test_alturas_override(self):
        cub_std = cubicar(RESULTADO_B1, altura_global=2.4)
        # Si BAÑO override a 2.2 → pintura debe ser menor
        cub_low = cubicar(RESULTADO_B1, altura_global=2.4, alturas_override={"COCINA": 2.2})
        partidas_std = {p["partida"]: p["cantidad"] for p in cub_std["partidas"]}
        partidas_low = {p["partida"]: p["cantidad"] for p in cub_low["partidas"]}
        assert partidas_low["pintura_interior_latex_2manos"] < partidas_std["pintura_interior_latex_2manos"]


# ─── cubicar_vial ─────────────────────────────────────────────────────────────

class TestCubicarVial:
    def _vial(self, nombre, area):
        return [{"nombre": nombre, "area_m2": area}]

    def test_calzada_genera_capas_pavimento(self):
        """Calzada → carpeta + base + sub-base + excavacion."""
        res = cubicar_vial(self._vial("CALZADA", 800.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas
        assert "base_granular_e200mm" in partidas
        assert "sub_base_granular_e200mm" in partidas
        assert "excavacion_tierra_comun" in partidas
        assert partidas["carpeta_asfaltica_e60mm"]["cantidad"] == pytest.approx(800.0, abs=0.1)
        assert partidas["carpeta_asfaltica_e60mm"]["tipo"] == "vial"

    def test_pista_genera_capas_pavimento(self):
        res = cubicar_vial(self._vial("PISTA 1", 400.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas

    def test_vereda_genera_acera_y_base(self):
        res = cubicar_vial(self._vial("VEREDA PONIENTE", 120.0))
        partidas = {p["partida"]: p for p in res}
        assert "acera_hormigon_e10cm" in partidas
        assert "base_granular_e200mm" in partidas
        assert partidas["acera_hormigon_e10cm"]["cantidad"] == pytest.approx(120.0, abs=0.1)

    def test_ciclovia_genera_partida(self):
        res = cubicar_vial(self._vial("CICLOVIA", 50.0))
        partidas = {p["partida"]: p for p in res}
        assert "ciclovia_pavimento" in partidas

    def test_cuneta_genera_ml(self):
        # area=10 m², ancho default 0.40 m → 25 ml
        res = cubicar_vial(self._vial("CUNETA REVESTIDA", 10.0))
        partidas = {p["partida"]: p for p in res}
        assert "cuneta_hormigon_revestida" in partidas
        assert partidas["cuneta_hormigon_revestida"]["unidad"] == "ml"
        assert partidas["cuneta_hormigon_revestida"]["cantidad"] == pytest.approx(25.0, abs=0.1)

    def test_tunel_genera_tres_partidas(self):
        """Tunel → excavacion m3 + revestimiento m2 + acero kg."""
        # area=200 m², default: ancho=8.5, alto=7.5
        # V_excav = 200 × 7.5 = 1500 m³
        res = cubicar_vial(self._vial("TUNEL", 200.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_tunel_roca" in partidas
        assert "revestimiento_tunel_hormigon" in partidas
        assert "acero_refuerzo_a630" in partidas
        assert partidas["excavacion_tunel_roca"]["unidad"] == "m3"
        assert partidas["excavacion_tunel_roca"]["cantidad"] == pytest.approx(1500.0, abs=5.0)

    def test_tunel_secciones_custom(self):
        """Secciones custom sobreescriben default."""
        sec = {"tunel": {"ancho_excav_m": 10.0, "alto_excav_m": 9.0, "espesor_revestimiento_m": 0.40}}
        res = cubicar_vial(self._vial("TUNEL", 100.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        # V = area × alto = 100 × 9 = 900 m³
        assert partidas["excavacion_tunel_roca"]["cantidad"] == pytest.approx(900.0, abs=1.0)

    def test_puente_genera_hormigon_acero_moldaje(self):
        """Puente → hormigon m3 + acero kg + moldaje m2."""
        # area=500 m², espesor default=0.55m → V=275 m³
        res = cubicar_vial(self._vial("PUENTE VEHICULAR", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "hormigon_armado_H30" in partidas
        assert "acero_refuerzo_a630" in partidas
        assert "moldaje_tablero" in partidas
        assert partidas["hormigon_armado_H30"]["cantidad"] == pytest.approx(275.0, abs=1.0)

    def test_muro_genera_precio_all_inclusive(self):
        """Muro → solo muro_contencion_hormigon m2 (all-inclusive, sin doble conteo)."""
        res = cubicar_vial(self._vial("MURO DE CONTENCION", 45.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas
        assert partidas["muro_contencion_hormigon"]["cantidad"] == pytest.approx(45.0, abs=0.1)
        # No componentes separados — evita doble conteo con precio all-inclusive
        assert "hormigon_armado_H30" not in partidas
        assert "moldaje_muro" not in partidas

    def test_galeria_usa_seccion_propia(self):
        """GALERIA usa sec['galeria'] (3.5×3.5) independiente de sec['tunel']."""
        res = cubicar_vial(self._vial("GALERIA SERVICIO", 49.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_tunel_roca" in partidas
        assert "revestimiento_tunel_hormigon" in partidas
        assert "acero_refuerzo_a630" in partidas
        # ancho default galeria=3.5 → longitud=49/3.5=14; V=14×3.5×3.5=171.5 m³
        assert partidas["excavacion_tunel_roca"]["cantidad"] == pytest.approx(171.5, abs=1.0)

    def test_galeria_custom_no_contamina_tunel(self):
        """Secciones custom de galeria no afectan la seccion de tunel."""
        sec = {"galeria": {"ancho_excav_m": 4.0, "alto_excav_m": 4.0,
                           "espesor_revestimiento_m": 0.30, "kg_acero_por_m3_revest": 70}}
        res_gal = cubicar_vial(self._vial("GALERIA BYPASS", 40.0), secciones=sec)
        res_tun = cubicar_vial(self._vial("TUNEL", 40.0), secciones=sec)
        pg = {p["partida"]: p for p in res_gal}
        pt = {p["partida"]: p for p in res_tun}
        # galeria V = (40/4)×4×4 = 160; tunel V = (40/8.5)×8.5×7.5 = 300
        assert pg["excavacion_tunel_roca"]["cantidad"] == pytest.approx(160.0, abs=1.0)
        assert pt["excavacion_tunel_roca"]["cantidad"] == pytest.approx(300.0, abs=1.0)

    def test_tunel_metro_seccion_circular(self):
        """TUNEL METRO usa seccion circular TBM (D=6.0m por default)."""
        res = cubicar_vial(self._vial("TUNEL METRO LINEA 3", 120.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_tunel_roca" in partidas
        assert "revestimiento_tunel_hormigon" in partidas
        assert "acero_refuerzo_a630" in partidas
        # D=6.0 → area_sec=π×3²=28.27 m²; longitud=120/6=20 m
        # V_excav = 20 × 28.27 ≈ 565.5 m³
        assert partidas["excavacion_tunel_roca"]["cantidad"] == pytest.approx(565.5, abs=2.0)

    def test_tunel_metro_custom_diametro(self):
        """Diametro configurable en secciones_civiles.yaml."""
        sec = {"tunel_metro": {"diametro_excav_m": 9.0, "espesor_dovelas_m": 0.35,
                               "kg_acero_por_m3_dovela": 110}}
        res = cubicar_vial(self._vial("TUNEL TBM PONIENTE", 180.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        # D=9.0 → area_sec=π×4.5²=63.62 m²; longitud=180/9=20 m
        # V_excav = 20 × 63.62 ≈ 1272.3 m³
        assert partidas["excavacion_tunel_roca"]["cantidad"] == pytest.approx(1272.3, abs=3.0)

    def test_descripciones_dinamicas_calzada(self):
        """Descripciones reflejan el espesor configurado, no un valor hardcodeado."""
        sec = {"pavimento_flexible": {"carpeta_asfaltica_m": 0.08,
                                       "base_granular_m": 0.25,
                                       "sub_base_granular_m": 0.25}}
        res = cubicar_vial(self._vial("CALZADA", 500.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert "80mm" in partidas["carpeta_asfaltica_e60mm"]["descripcion"]
        assert "250mm" in partidas["base_granular_e200mm"]["descripcion"]
        assert "250mm" in partidas["sub_base_granular_e200mm"]["descripcion"]

    def test_elemento_sin_categoria_ignorado(self):
        # MEDIANA es vial pero sin cat mapeada → lista vacia
        res = cubicar_vial(self._vial("MEDIANA", 20.0))
        assert res == []

    def test_multiples_calzadas_acumulan(self):
        viales = [
            {"nombre": "CALZADA PONIENTE", "area_m2": 400.0},
            {"nombre": "CALZADA ORIENTE", "area_m2": 400.0},
        ]
        res = cubicar_vial(viales)
        partidas = {p["partida"]: p for p in res}
        assert partidas["carpeta_asfaltica_e60mm"]["cantidad"] == pytest.approx(800.0, abs=0.1)

    def test_tipo_vial_en_todas_las_partidas(self):
        viales = [
            {"nombre": "CALZADA", "area_m2": 200.0},
            {"nombre": "TUNEL", "area_m2": 100.0},
        ]
        res = cubicar_vial(viales)
        assert all(p["tipo"] == "vial" for p in res)

    def test_supuesto_documenta_dimension(self):
        res = cubicar_vial(self._vial("CUNETA", 8.0))
        partidas = {p["partida"]: p for p in res}
        assert partidas["cuneta_hormigon_revestida"]["supuesto"] != ""

    # ── Bug fix: MIRADOR no debe clasificarse como calzada ────────────────────

    def test_mirador_calzada_ignorado(self):
        """MIRADOR CALZADA KMxxx no genera partidas de pavimento."""
        res = cubicar_vial(self._vial("MIRADOR CALZADA KM210", 480.0))
        assert res == [], "MIRADOR no debe generar partidas — contiene CALZADA pero es obra menor"

    def test_paradero_ignorado(self):
        res = cubicar_vial(self._vial("PARADERO BUS", 20.0))
        assert res == []

    # ── Nuevas categorías viales ──────────────────────────────────────────────

    def test_terraplen_genera_m3(self):
        """TERRAPLEN → m³ = area × altura_media (default 2.0m)."""
        res = cubicar_vial(self._vial("TERRAPLEN COMPACTADO", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "terraplen_compactado" in partidas
        assert partidas["terraplen_compactado"]["unidad"] == "m3"
        assert partidas["terraplen_compactado"]["cantidad"] == pytest.approx(2000.0, abs=0.1)

    def test_terraplen_altura_custom(self):
        sec = {"terraplen": {"altura_media_m": 3.0}}
        res = cubicar_vial(self._vial("TERRAPLEN", 500.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert partidas["terraplen_compactado"]["cantidad"] == pytest.approx(1500.0, abs=0.1)

    def test_alcantarilla_genera_ml(self):
        """ALCANTARILLA → ml = area / ancho_interno (default 1.5m)."""
        res = cubicar_vial(self._vial("ALCANTARILLA MARCO HORMIGON", 15.0))
        partidas = {p["partida"]: p for p in res}
        assert "alcantarilla_marco_hormigon" in partidas
        assert partidas["alcantarilla_marco_hormigon"]["unidad"] == "ml"
        assert partidas["alcantarilla_marco_hormigon"]["cantidad"] == pytest.approx(10.0, abs=0.1)

    def test_demarcacion_genera_m2(self):
        res = cubicar_vial(self._vial("DEMARCACION VIAL", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "demarcacion_vial_termoplastico" in partidas
        assert partidas["demarcacion_vial_termoplastico"]["unidad"] == "m2"
        assert partidas["demarcacion_vial_termoplastico"]["cantidad"] == pytest.approx(500.0)

    def test_senaletiva_genera_unidades(self):
        """SENALETIVA → un = area / 1.5 m² (default)."""
        res = cubicar_vial(self._vial("SENALETIVA VERTICAL KM0-21", 15.0))
        partidas = {p["partida"]: p for p in res}
        assert "senal_vial_retrorreflectante" in partidas
        assert partidas["senal_vial_retrorreflectante"]["unidad"] == "un"
        # 15 / 1.5 = 10 señales
        assert partidas["senal_vial_retrorreflectante"]["cantidad"] == pytest.approx(10.0, abs=0.5)

    def test_iluminacion_genera_unidades(self):
        """ILUMINACION → un = area / 4.0 m² (default)."""
        res = cubicar_vial(self._vial("ILUMINACION VIAL", 100.0))
        partidas = {p["partida"]: p for p in res}
        assert "luminaria_led_vial" in partidas
        assert partidas["luminaria_led_vial"]["unidad"] == "un"
        # 100 / 4 = 25 postes
        assert partidas["luminaria_led_vial"]["cantidad"] == pytest.approx(25.0, abs=0.5)

    def test_guardavia_genera_ml(self):
        """GUARDAVIA → ml = area / 0.5m (default)."""
        res = cubicar_vial(self._vial("GUARDAVIA FLEXIBLE W", 50.0))
        partidas = {p["partida"]: p for p in res}
        assert "guardavia_flexible_w" in partidas
        assert partidas["guardavia_flexible_w"]["unidad"] == "ml"
        # 50 / 0.5 = 100 ml
        assert partidas["guardavia_flexible_w"]["cantidad"] == pytest.approx(100.0, abs=0.1)

    def test_revegetacion_genera_m2(self):
        res = cubicar_vial(self._vial("REVEGETACION TALUDES", 200.0))
        partidas = {p["partida"]: p for p in res}
        assert "revegetacion_hidrosiembra" in partidas
        assert partidas["revegetacion_hidrosiembra"]["unidad"] == "m2"
        assert partidas["revegetacion_hidrosiembra"]["cantidad"] == pytest.approx(200.0)

    # ── Subrasante en calzada (nueva capa) ────────────────────────────────────

    def test_calzada_incluye_subrasante_por_defecto(self):
        """Default subrasante_m=0.15 genera partida subrasante_estabilizada."""
        res = cubicar_vial(self._vial("CALZADA", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "subrasante_estabilizada" in partidas
        assert partidas["subrasante_estabilizada"]["unidad"] == "m2"
        assert partidas["subrasante_estabilizada"]["cantidad"] == pytest.approx(1000.0)

    def test_calzada_excav_incluye_subrasante(self):
        """Excavacion = area × (cap+base+sub+subrasante) = 1000 × 0.61 = 610 m³."""
        res = cubicar_vial(self._vial("CALZADA", 1000.0))
        partidas = {p["partida"]: p for p in res}
        # 0.06+0.20+0.20+0.15 = 0.61 m
        assert partidas["excavacion_tierra_comun"]["cantidad"] == pytest.approx(610.0, abs=1.0)

    def test_calzada_sin_subrasante_cuando_cero(self):
        """subrasante_m=0.0 omite la partida y no suma a excavacion."""
        sec = {"pavimento_flexible": {"subrasante_m": 0.0}}
        res = cubicar_vial(self._vial("CALZADA", 1000.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert "subrasante_estabilizada" not in partidas
        # excavacion = 1000 × 0.46 = 460
        assert partidas["excavacion_tierra_comun"]["cantidad"] == pytest.approx(460.0, abs=1.0)

    # ── Corte y escarpe ───────────────────────────────────────────────────────

    def test_corte_genera_m3(self):
        """CORTE EN ROCA → m³ = area × profundidad_media (default 3.0m)."""
        res = cubicar_vial(self._vial("CORTE EN ROCA", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_en_corte" in partidas
        assert partidas["excavacion_en_corte"]["unidad"] == "m3"
        assert partidas["excavacion_en_corte"]["cantidad"] == pytest.approx(1500.0, abs=0.1)

    def test_corte_profundidad_custom(self):
        sec = {"corte": {"profundidad_media_m": 5.0}}
        res = cubicar_vial(self._vial("EXCAVACION EN CORTE", 200.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert partidas["excavacion_en_corte"]["cantidad"] == pytest.approx(1000.0, abs=0.1)

    def test_escarpe_genera_m2(self):
        res = cubicar_vial(self._vial("ESCARPE LIMPIEZA DE TERRENO", 800.0))
        partidas = {p["partida"]: p for p in res}
        assert "escarpe_y_limpieza" in partidas
        assert partidas["escarpe_y_limpieza"]["unidad"] == "m2"
        assert partidas["escarpe_y_limpieza"]["cantidad"] == pytest.approx(800.0)

    # ── Canal revestido ───────────────────────────────────────────────────────

    def test_canal_genera_ml(self):
        """CANAL → ml = area / ancho_canal (default 2.0m)."""
        res = cubicar_vial(self._vial("CANAL REVESTIDO", 100.0))
        partidas = {p["partida"]: p for p in res}
        assert "canal_hormigon_revestido" in partidas
        assert partidas["canal_hormigon_revestido"]["unidad"] == "ml"
        assert partidas["canal_hormigon_revestido"]["cantidad"] == pytest.approx(50.0, abs=0.1)

    # ── Pasarela peatonal ─────────────────────────────────────────────────────

    def test_pasarela_genera_tres_partidas(self):
        """PASARELA → hormigon m³ + acero kg + moldaje m²."""
        # area=120 m², espesor default=0.25m → V=30 m³
        res = cubicar_vial(self._vial("PASARELA PEATONAL", 120.0))
        partidas = {p["partida"]: p for p in res}
        assert "hormigon_armado_H30" in partidas
        assert "acero_refuerzo_a630" in partidas
        assert "moldaje_tablero" in partidas
        assert partidas["hormigon_armado_H30"]["cantidad"] == pytest.approx(30.0, abs=0.5)

    # ── Adoquín ───────────────────────────────────────────────────────────────

    def test_adoquin_genera_m2(self):
        res = cubicar_vial(self._vial("ADOQUIN HORMIGON", 300.0))
        partidas = {p["partida"]: p for p in res}
        assert "pavimento_adoquin_hormigon" in partidas
        assert partidas["pavimento_adoquin_hormigon"]["cantidad"] == pytest.approx(300.0)

    # ── Rotonda (alias calzada) ───────────────────────────────────────────────

    def test_rotonda_genera_capas_pavimento(self):
        """ROTONDA se clasifica como calzada y genera mismas capas."""
        res = cubicar_vial(self._vial("ROTONDA KM5", 600.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas
        assert partidas["carpeta_asfaltica_e60mm"]["cantidad"] == pytest.approx(600.0)

    # ── Zapata muro (opcional) ────────────────────────────────────────────────

    def test_muro_solo_m2_all_inclusive(self):
        """Muro usa precio all-inclusive: solo muro_contencion_hormigon, sin hormigon separado."""
        res = cubicar_vial(self._vial("MURO DE CONTENCION", 100.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas
        assert partidas["muro_contencion_hormigon"]["cantidad"] == pytest.approx(100.0, abs=0.1)
        assert "hormigon_armado_H30" not in partidas

    def test_muro_con_geotextil(self):
        """geotextil_factor=1.1 agrega partida geotextil_drenaje."""
        sec = {"muro_contencion": {"espesor_m": 0.40, "kg_acero_por_m3": 80,
                                    "zapata_factor": 0.0, "geotextil_factor": 1.1}}
        res = cubicar_vial(self._vial("MURO DE CONTENCION", 100.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert "geotextil_drenaje" in partidas
        assert partidas["geotextil_drenaje"]["cantidad"] == pytest.approx(110.0, abs=0.1)

    # ── Puente con estribos ───────────────────────────────────────────────────

    def test_puente_sin_estribo_por_defecto(self):
        """estribo_factor=0 por defecto: hormigon solo del tablero."""
        res = cubicar_vial(self._vial("PUENTE VEHICULAR", 500.0))
        partidas = {p["partida"]: p for p in res}
        # solo tablero: 500×0.55=275 m³
        assert partidas["hormigon_armado_H30"]["cantidad"] == pytest.approx(275.0, abs=1.0)

    def test_puente_con_estribo(self):
        """estribo_factor=0.35 agrega 35% de vol_tablero como estribos."""
        sec = {"puente": {"espesor_tablero_m": 0.55, "kg_acero_por_m3_tablero": 120,
                           "estribo_factor": 0.35}}
        res = cubicar_vial(self._vial("PUENTE VEHICULAR", 500.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        # tablero=275 + estribos=275×0.35=96.25 → total=371.25 m³
        assert partidas["hormigon_armado_H30"]["cantidad"] == pytest.approx(371.25, abs=1.0)

    # ── Barrera NJ ────────────────────────────────────────────────────────────

    def test_barrera_nj_genera_ml(self):
        """BARRERA HORMIGON → ml = area / 0.60m (default)."""
        res = cubicar_vial(self._vial("BARRERA HORMIGON NEW JERSEY", 60.0))
        partidas = {p["partida"]: p for p in res}
        assert "barrera_hormigon_nj" in partidas
        assert partidas["barrera_hormigon_nj"]["unidad"] == "ml"
        # 60 / 0.60 = 100 ml
        assert partidas["barrera_hormigon_nj"]["cantidad"] == pytest.approx(100.0, abs=0.1)

    def test_barrera_nj_no_confunde_con_guardavia(self):
        """BARRERA HORMIGON → barrera_nj, BARRERA METALICA → guardavia."""
        res_nj = cubicar_vial(self._vial("BARRERA HORMIGON NJ", 60.0))
        res_gv = cubicar_vial(self._vial("BARRERA METALICA W", 60.0))
        partidas_nj = {p["partida"]: p for p in res_nj}
        partidas_gv = {p["partida"]: p for p in res_gv}
        assert "barrera_hormigon_nj" in partidas_nj
        assert "guardavia_flexible_w" in partidas_gv
        assert "barrera_hormigon_nj" not in partidas_gv

    # ── Pozo de inspección ────────────────────────────────────────────────────

    def test_pozo_inspeccion_genera_unidades(self):
        """POZO DE INSPECCION → un = area / 4.0 m² (default)."""
        res = cubicar_vial(self._vial("POZO DE INSPECCION", 20.0))
        partidas = {p["partida"]: p for p in res}
        assert "pozo_inspeccion_hormigon" in partidas
        assert partidas["pozo_inspeccion_hormigon"]["unidad"] == "un"
        # 20 / 4 = 5 pozos
        assert partidas["pozo_inspeccion_hormigon"]["cantidad"] == pytest.approx(5.0, abs=0.5)

    # ── Mejoramiento de suelo ─────────────────────────────────────────────────

    def test_mejoramiento_suelo_genera_m2(self):
        res = cubicar_vial(self._vial("MEJORAMIENTO SUBRASANTE CAL", 1200.0))
        partidas = {p["partida"]: p for p in res}
        assert "mejoramiento_suelo_cal" in partidas
        assert partidas["mejoramiento_suelo_cal"]["unidad"] == "m2"
        assert partidas["mejoramiento_suelo_cal"]["cantidad"] == pytest.approx(1200.0)

    def test_colector_pvc_genera_ml(self):
        # 160 m² area planta, zanja 0.8m → 200 ml
        res = cubicar_vial(self._vial("COLECTOR PVC KM5", 160.0))
        partidas = {p["partida"]: p for p in res}
        assert "colector_pvc_300mm" in partidas
        assert partidas["colector_pvc_300mm"]["unidad"] == "ml"
        assert partidas["colector_pvc_300mm"]["cantidad"] == pytest.approx(200.0)

    def test_colector_hormigon_genera_ml(self):
        # keyword HORMIGON → colector_hormigon_600mm
        res = cubicar_vial(self._vial("COLECTOR HORMIGON KM8", 240.0))
        partidas = {p["partida"]: p for p in res}
        assert "colector_hormigon_600mm" in partidas
        assert "colector_pvc_300mm" not in partidas

    def test_berma_genera_capas_pavimento(self):
        # BERMA should be treated as calzada (flexible pavement)
        res = cubicar_vial(self._vial("BERMA PAVIMENTADA LATERAL", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas
        assert "base_granular_e200mm" in partidas

    def test_paso_superior_vehicular_genera_puente(self):
        res = cubicar_vial(self._vial("PASO SUPERIOR VEHICULAR KM14", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "hormigon_armado_H30" in partidas
        assert "moldaje_tablero" in partidas

    def test_paso_bajo_nivel_genera_puente(self):
        res = cubicar_vial(self._vial("PASO BAJO NIVEL ACCESO", 300.0))
        partidas = {p["partida"]: p for p in res}
        assert "hormigon_armado_H30" in partidas

    def test_losa_aproximacion_genera_puente(self):
        res = cubicar_vial(self._vial("LOSA DE APROXIMACION PUENTE KM8", 200.0))
        partidas = {p["partida"]: p for p in res}
        assert "hormigon_armado_H30" in partidas

    def test_ensanche_genera_calzada(self):
        res = cubicar_vial(self._vial("ENSANCHE CALZADA KM3", 800.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas

    def test_excavacion_masiva_genera_corte(self):
        res = cubicar_vial(self._vial("EXCAVACION MASIVA SECTOR A", 600.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_en_corte" in partidas

    def test_relleno_estructural_genera_terraplen(self):
        res = cubicar_vial(self._vial("RELLENO ESTRUCTURAL ESTRIBO", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "terraplen_compactado" in partidas

    def test_mediana_verde_genera_revegetacion(self):
        res = cubicar_vial(self._vial("MEDIANA VERDE SEPARADOR CENTRAL", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "revegetacion_hidrosiembra" in partidas

    def test_mediana_central_sin_qualifier_genera_revegetacion(self):
        # MEDIANA sin tipo → default revegetacion (pasto/arbustos)
        res = cubicar_vial(self._vial("MEDIANA CENTRAL KM5", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "revegetacion_hidrosiembra" in partidas

    def test_acentos_tunel_normalizado(self):
        # "TÚNEL" con acento debe matchear igual que "TUNEL"
        res = cubicar_vial(self._vial("TÚNEL GALLEGUILLOS", 1620.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_tunel_roca" in partidas

    def test_acentos_cuneta_normalizada(self):
        res = cubicar_vial(self._vial("CUNETA REVESTIDA HORMIGÓN", 600.0))
        partidas = {p["partida"]: p for p in res}
        assert "cuneta_hormigon_revestida" in partidas

    def test_recarpeteo_solo_carpeta_sin_excavacion(self):
        # Conservacion: solo carpeta asfaltica, sin base ni excavacion
        res = cubicar_vial(self._vial("RECARPETEO ASFALTICO KM5", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "carpeta_asfaltica_e60mm" in partidas
        assert "base_granular_e200mm" not in partidas, "recarpeteo no repone base existente"
        assert "excavacion_tierra_comun" not in partidas, "recarpeteo no excava"

    def test_afirmado_granular_sin_carpeta(self):
        # Camino ripio: base + sub_base + excavacion, SIN carpeta asfaltica
        res = cubicar_vial(self._vial("AFIRMADO GRANULAR CAMINO RURAL", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "base_granular_e200mm" in partidas
        assert "sub_base_granular_e200mm" in partidas
        assert "carpeta_asfaltica_e60mm" not in partidas, "camino granular no lleva carpeta asfaltica"

    def test_camino_ripio_genera_excavacion(self):
        # Excavacion = area × (base + sub_base) = 1000 × 0.40 = 400 m³
        res = cubicar_vial(self._vial("CAMINO RIPIO ACCESO", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_tierra_comun" in partidas
        assert partidas["excavacion_tierra_comun"]["cantidad"] == pytest.approx(400.0)

    def test_trinchera_genera_corte(self):
        res = cubicar_vial(self._vial("TRINCHERA COLECTOR KM8", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "excavacion_en_corte" in partidas

    def test_geotextil_standalone_genera_m2(self):
        res = cubicar_vial(self._vial("GEOTEXTIL SEPARACION BASE GRANULAR", 2000.0))
        partidas = {p["partida"]: p for p in res}
        assert "geotextil_drenaje" in partidas
        assert partidas["geotextil_drenaje"]["cantidad"] == pytest.approx(2000.0)

    def test_geomalla_genera_geotextil(self):
        res = cubicar_vial(self._vial("GEOMALLA BIAXIAL TALUD", 1500.0))
        partidas = {p["partida"]: p for p in res}
        assert "geotextil_drenaje" in partidas

    def test_mediana_pavimentada_genera_adoquin(self):
        res = cubicar_vial(self._vial("MEDIANA PAVIMENTADA ADOQUIN", 400.0))
        partidas = {p["partida"]: p for p in res}
        assert "pavimento_adoquin_hormigon" in partidas

    # ── Ciclovia (fix: incluye base granular) ────────────────────────────────

    def test_ciclovia_genera_pavimento_y_base(self):
        """CICLOVIA → ciclovia_pavimento m² + base_granular m² (e=150mm)."""
        res = cubicar_vial(self._vial("CICLOVIA SEGREGADA", 560.0))
        partidas = {p["partida"]: p for p in res}
        assert "ciclovia_pavimento" in partidas
        assert "base_granular_e200mm" in partidas
        assert partidas["ciclovia_pavimento"]["cantidad"] == pytest.approx(560.0)
        assert partidas["base_granular_e200mm"]["cantidad"] == pytest.approx(560.0)

    def test_ciclopista_genera_ciclovia(self):
        res = cubicar_vial(self._vial("CICLOPISTA KM3", 400.0))
        partidas = {p["partida"]: p for p in res}
        assert "ciclovia_pavimento" in partidas
        assert "base_granular_e200mm" in partidas

    def test_via_ciclista_genera_ciclovia(self):
        res = cubicar_vial(self._vial("VIA CICLISTA NORTE", 300.0))
        partidas = {p["partida"]: p for p in res}
        assert "ciclovia_pavimento" in partidas

    def test_ciclovia_base_granular_override(self):
        """ciclovia.base_granular_m override via secciones."""
        sec = {"ciclovia": {"base_granular_m": 0.10}}
        res = cubicar_vial(self._vial("CICLOVIA SEGREGADA", 1000.0), secciones=sec)
        partidas = {p["partida"]: p for p in res}
        assert "base_granular_e200mm" in partidas
        assert partidas["base_granular_e200mm"]["cantidad"] == pytest.approx(1000.0)

    # ── Calzada rigida (fix: incluye excavacion) ──────────────────────────────

    def test_calzada_rigida_genera_tres_partidas(self):
        """PAVIMENTO RIGIDO → pavimento_hormigon_rigido + sub_base + excavacion."""
        res = cubicar_vial(self._vial("PAVIMENTO RIGIDO H30", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert "pavimento_hormigon_rigido" in partidas
        assert "sub_base_granular_e200mm" in partidas
        assert "excavacion_tierra_comun" in partidas

    def test_calzada_rigida_excavacion_volumen(self):
        """Excavacion = area × (losa + sub_base) = 1000 × (0.22+0.15) = 370 m³."""
        res = cubicar_vial(self._vial("PAVIMENTO RIGIDO H30", 1000.0))
        partidas = {p["partida"]: p for p in res}
        assert partidas["excavacion_tierra_comun"]["cantidad"] == pytest.approx(370.0, abs=1.0)

    def test_calzada_hormigon_genera_rigida(self):
        res = cubicar_vial(self._vial("CALZADA HORMIGON KM6", 800.0))
        partidas = {p["partida"]: p for p in res}
        assert "pavimento_hormigon_rigido" in partidas

    # ── Nuevas keywords calzada_granular ─────────────────────────────────────

    def test_espaldon_genera_calzada_granular(self):
        """ESPALDON (berma no pavimentada) → calzada_granular sin carpeta asfaltica."""
        res = cubicar_vial(self._vial("ESPALDON GRANULAR", 600.0))
        partidas = {p["partida"]: p for p in res}
        assert "base_granular_e200mm" in partidas
        assert "sub_base_granular_e200mm" in partidas
        assert "carpeta_asfaltica_e60mm" not in partidas

    def test_camino_lateral_genera_calzada_granular(self):
        res = cubicar_vial(self._vial("CAMINO LATERAL DE SERVICIO", 500.0))
        partidas = {p["partida"]: p for p in res}
        assert "base_granular_e200mm" in partidas
        assert "carpeta_asfaltica_e60mm" not in partidas

    # ── Nuevas keywords colector ──────────────────────────────────────────────

    def test_tuberia_hdpe_genera_colector(self):
        """TUBERIA HDPE → colector (no hormigon keyword → PVC)."""
        res = cubicar_vial(self._vial("TUBERIA HDPE DRENAJE PLUVIAL", 160.0))
        partidas = {p["partida"]: p for p in res}
        assert "colector_pvc_300mm" in partidas

    def test_tuberia_corrugada_genera_colector(self):
        res = cubicar_vial(self._vial("TUBERIA CORRUGADA DRENAJE KM4", 80.0))
        partidas = {p["partida"]: p for p in res}
        assert "colector_pvc_300mm" in partidas
        assert partidas["colector_pvc_300mm"]["unidad"] == "ml"

    def test_subcolector_genera_colector(self):
        res = cubicar_vial(self._vial("SUBCOLECTOR PVC 200MM", 40.0))
        partidas = {p["partida"]: p for p in res}
        assert "colector_pvc_300mm" in partidas

    # ── Nuevas keywords muro ──────────────────────────────────────────────────

    def test_enrocado_genera_muro(self):
        """ENROCADO → muro_contencion_hormigon (all-inclusive, precio escollera)."""
        res = cubicar_vial(self._vial("ENROCADO TALUD RIO", 800.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas
        assert partidas["muro_contencion_hormigon"]["cantidad"] == pytest.approx(800.0)

    def test_enrocamiento_genera_muro(self):
        res = cubicar_vial(self._vial("ENROCAMIENTO RIBERA CANAL", 400.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas

    def test_muro_gavion_genera_muro(self):
        res = cubicar_vial(self._vial("MURO GAVION METALICO", 300.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas

    def test_muro_tierra_armada_genera_muro(self):
        res = cubicar_vial(self._vial("MURO TIERRA ARMADA KM12", 1500.0))
        partidas = {p["partida"]: p for p in res}
        assert "muro_contencion_hormigon" in partidas

    def test_cubicar_integra_vial_en_partidas(self):
        """cubicar() incluye partidas viales junto a las residenciales."""
        res = dict(RESULTADO_B1)
        res["recintos"] = RECINTOS_B1 + [
            {"nombre": "CALZADA", "area_m2": 500.0, "confianza": 0.9, "departamento": None},
            {"nombre": "VEREDA", "area_m2": 100.0, "confianza": 0.9, "departamento": None},
        ]
        cub = cubicar(res, altura_global=2.4)
        partidas = {p["partida"]: p for p in cub["partidas"]}
        assert "carpeta_asfaltica_e60mm" in partidas
        assert "acera_hormigon_e10cm" in partidas
        # Residencial no se contamina
        assert partidas["cielo_volcanita_pintado"]["cantidad"] == pytest.approx(58.3, abs=0.1)


# ─── presupuesto ──────────────────────────────────────────────────────────────

@pytest.mark.skipif(not PRECIOS_CSV.exists(), reason="precios_cl.csv no encontrado")
class TestPresupuesto:
    @pytest.fixture(scope="class")
    def precios(self):
        return cargar_precios(PRECIOS_CSV)

    @pytest.fixture(scope="class")
    def cubicacion(self):
        return cubicar(RESULTADO_B1, altura_global=2.4)

    @pytest.fixture(scope="class")
    def pres(self, cubicacion, precios):
        return presupuestar(cubicacion, precios)

    def test_precios_cargados(self, precios):
        assert "pintura_interior_latex_2manos" in precios
        assert precios["pintura_interior_latex_2manos"]["precio_clp"] == 3500

    def test_sin_faltantes(self, pres):
        assert pres["faltantes"] == [], f"Partidas sin precio: {pres['faltantes']}"

    def test_total_positivo(self, pres):
        assert pres["subtotales"]["total_con_iva"] > 0

    def test_iva_separado(self, pres):
        s = pres["subtotales"]
        # IVA como línea separada: neto × (1 + GGU) × IVA
        assert s["iva_monto"] == int(round(s["con_gg"] * s["iva_pct"]))

    def test_total_igual_con_gg_mas_iva(self, pres):
        s = pres["subtotales"]
        assert s["total_con_iva"] == s["con_gg"] + s["iva_monto"]

    def test_estructura_lineas(self, pres):
        for l in pres["lineas"]:
            assert "partida" in l
            assert "cantidad" in l
            assert "precio_unitario_clp" in l
            assert l["subtotal_clp"] == int(round(l["cantidad"] * l["precio_unitario_clp"]))


# ─── excel (validacion programatica) ──────────────────────────────────────────

@pytest.mark.skipif(not PRECIOS_CSV.exists(), reason="precios_cl.csv no encontrado")
class TestExcel:
    @pytest.fixture(scope="class")
    def excel_path(self, tmp_path_factory):
        tmp = tmp_path_factory.mktemp("excel")
        ruta = tmp / "test_b1.xlsx"
        cub = cubicar(RESULTADO_B1, altura_global=2.4)
        precios = cargar_precios(PRECIOS_CSV)
        pres = presupuestar(cub, precios)
        exportar_excel(RESULTADO_B1, None, cub, pres, ruta)
        return ruta

    def test_4_hojas(self, excel_path):
        wb = load_workbook(excel_path)
        assert set(wb.sheetnames) == {"Resumen", "Recintos", "Cubicacion", "Trazabilidad"}

    def test_cubicacion_tiene_formulas_subtotal(self, excel_path):
        wb = load_workbook(excel_path)
        ws = wb["Cubicacion"]
        formulas = [
            ws.cell(row=r, column=6).value
            for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=6).value, str)
            and ws.cell(row=r, column=6).value.startswith("=")
        ]
        assert len(formulas) >= 2, "Debe haber al menos 1 fórmula de subtotal + SUM"

    def test_total_es_formula(self, excel_path):
        wb = load_workbook(excel_path)
        ws = wb["Cubicacion"]
        # La última fórmula en col F con "+" debe ser el total
        formulas = [
            ws.cell(row=r, column=6).value
            for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=6).value, str)
            and ws.cell(row=r, column=6).value.startswith("=")
        ]
        assert any("+" in f for f in formulas), "Total con IVA debe ser =conGG+iva"

    def test_hoja_resumen_tiene_total(self, excel_path):
        wb = load_workbook(excel_path)
        ws = wb["Resumen"]
        valores = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("TOTAL" in str(v) for v in valores if v)

    def test_hoja_recintos_tiene_6_recintos(self, excel_path):
        wb = load_workbook(excel_path)
        ws = wb["Recintos"]
        nombres = [
            ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=2).value
        ]
        assert len(nombres) >= 6
