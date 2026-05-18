#!/usr/bin/env python3
"""
scraper_precios.py — descarga y extrae precios de construccion Chile

EJECUTAR EN TU MAQUINA LOCAL (no en cloud — red bloqueada).

Instalacion previa:
    pip install playwright curl_cffi openpyxl
    python -m playwright install chromium

Uso:
    python scraper_precios.py --fuente minvu
    python scraper_precios.py --fuente cype
    python scraper_precios.py --fuente ondac          # requiere ONDAC_USER + ONDAC_PASS
    python scraper_precios.py --fuente all
    python scraper_precios.py --fuente minvu --output staging.csv
    python scraper_precios.py --diff staging.csv      # muestra partidas nuevas vs precios_cl.csv

Salida: precios_staging.csv  (revisar antes de merge a precios_cl.csv)
"""

import argparse
import csv
import io
import os
import re
import sys
import time
from pathlib import Path

# ─── helpers ──────────────────────────────────────────────────────────────────

def _slug(texto: str) -> str:
    """Convierte texto libre a clave snake_case para partida."""
    t = texto.lower().strip()
    t = re.sub(r"[áàä]", "a", t)
    t = re.sub(r"[éèë]", "e", t)
    t = re.sub(r"[íìï]", "i", t)
    t = re.sub(r"[óòö]", "o", t)
    t = re.sub(r"[úùü]", "u", t)
    t = re.sub(r"[ñ]", "n", t)
    t = re.sub(r"[^a-z0-9]+", "_", t)
    t = t.strip("_")
    return t


def _clp(valor) -> int:
    """Convierte valor (str con puntos/comas o float) a int CLP."""
    if valor is None:
        return 0
    s = str(valor).replace(".", "").replace(",", "").replace("$", "").strip()
    try:
        return int(float(s))
    except ValueError:
        return 0


def _guardar_csv(partidas: list[dict], path: Path) -> None:
    campos = ["partida", "unidad", "precio_clp", "fuente", "fecha"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        w.writeheader()
        w.writerows(partidas)
    print(f"[OK] {len(partidas)} partidas → {path}")


# ─── MINVU ────────────────────────────────────────────────────────────────────

MINVU_XLSX_URL = (
    "https://www.minvu.gob.cl/wp-content/uploads/"
    "2025/03/Analisis-de-precios-unitarios-tipo-v2025.xlsx"
)
MINVU_PDF_URL = (
    "https://www.minvu.gob.cl/wp-content/uploads/"
    "2025/03/TABLA-DE-PRECIOS-REFERENCIALES-DS-27-REGION-METROPOLITANA.pdf"
)


def _descargar_minvu_xlsx() -> bytes | None:
    """Descarga el Excel de APU MINVU con impersonacion Chrome."""
    try:
        from curl_cffi import requests as cf
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet,*/*"
            ),
            "Accept-Language": "es-CL,es;q=0.9,en;q=0.7",
            "Referer": "https://www.minvu.gob.cl/",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
        }
        r = cf.get(MINVU_XLSX_URL, headers=headers, impersonate="chrome124", timeout=30)
        if r.status_code == 200 and len(r.content) > 5000:
            print(f"  [MINVU] xlsx descargado: {len(r.content):,} bytes")
            return r.content
        print(f"  [MINVU] xlsx: HTTP {r.status_code} — intentando con Playwright...")
        return None
    except ImportError:
        print("  [MINVU] curl_cffi no disponible")
        return None


def _descargar_minvu_playwright(url: str) -> bytes | None:
    """Descarga un archivo MINVU usando Playwright (bypass anti-bot)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [MINVU] playwright no instalado: pip install playwright && python -m playwright install chromium")
        return None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            accept_downloads=True,
        )
        page = ctx.new_page()
        # Visitar home primero para establecer cookies
        page.goto("https://www.minvu.gob.cl/", timeout=30000)
        time.sleep(2)

        # Descargar archivo
        with page.expect_download(timeout=30000) as dl:
            page.goto(url)
        download = dl.value
        tmp = Path("/tmp") / download.suggested_filename
        download.save_as(tmp)
        data = tmp.read_bytes()
        browser.close()
        print(f"  [MINVU] descargado via Playwright: {len(data):,} bytes → {tmp}")
        return data


def _parsear_minvu_xlsx(data: bytes) -> list[dict]:
    """Extrae partidas del Excel APU MINVU."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [MINVU] openpyxl no instalado")
        return []

    wb = load_workbook(io.BytesIO(data), data_only=True)
    partidas = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  [MINVU] hoja: {sheet_name} ({ws.max_row} filas)")

        # Detectar cabecera buscando columnas con "descripcion", "unidad", "precio"
        header_row = None
        col_desc = col_unidad = col_precio = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                cv = str(cell).lower()
                if any(k in cv for k in ["descripcion", "descripción", "partida", "item"]):
                    col_desc = i
                if any(k in cv for k in ["unidad", "und", "ud"]):
                    col_unidad = i
                if any(k in cv for k in ["precio", "total", "costo", "valor"]):
                    col_precio = i
            if col_desc is not None and col_precio is not None:
                header_row = ws._current_row
                break

        if col_desc is None:
            # Asumir columnas por posicion: A=desc, B=unidad, C=precio
            col_desc, col_unidad, col_precio = 0, 1, 2

        for row in ws.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
            if not row or row[col_desc] is None:
                continue
            desc = str(row[col_desc]).strip()
            if len(desc) < 4 or desc.lower() in ("descripcion", "partida", "item"):
                continue
            unidad = str(row[col_unidad]).strip() if col_unidad is not None and row[col_unidad] else "m2"
            precio = _clp(row[col_precio]) if col_precio is not None else 0
            if precio <= 0:
                continue
            partidas.append({
                "partida": _slug(desc),
                "unidad": unidad.lower(),
                "precio_clp": precio,
                "fuente": "MINVU APU v2025",
                "fecha": "2025-03",
                "_descripcion_original": desc,
            })

    print(f"  [MINVU] {len(partidas)} partidas extraidas del xlsx")
    return partidas


def _parsear_minvu_pdf(data: bytes) -> list[dict]:
    """Extrae tabla de precios referenciales del PDF MINVU (DS27 RM)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  [MINVU] pymupdf no instalado")
        return []

    doc = fitz.open(stream=data, filetype="pdf")
    partidas = []
    # Patron: descripcion + unidad + precio UF + precio CLP
    patron = re.compile(
        r"(.{10,60}?)\s+(m2|m3|ml|kg|un|gl|lm)\s+([\d.,]+)\s*(?:UF)?\s+([\d.,]+)",
        re.IGNORECASE,
    )
    for page in doc:
        text = page.get_text()
        for m in patron.finditer(text):
            desc, und, _uf, clp_str = m.groups()
            precio = _clp(clp_str)
            if precio < 500:
                continue
            partidas.append({
                "partida": _slug(desc.strip()),
                "unidad": und.lower(),
                "precio_clp": precio,
                "fuente": "MINVU DS27 RM 2025",
                "fecha": "2025-03",
                "_descripcion_original": desc.strip(),
            })

    print(f"  [MINVU] {len(partidas)} partidas extraidas del PDF")
    return partidas


def scrape_minvu() -> list[dict]:
    print("[MINVU] Iniciando scraping...")

    # 1. Intentar Excel (mas rico)
    data = _descargar_minvu_xlsx()
    if data is None:
        data = _descargar_minvu_playwright(MINVU_XLSX_URL)
    if data:
        return _parsear_minvu_xlsx(data)

    # 2. Fallback: PDF DS27
    print("[MINVU] Intentando PDF DS27 RM...")
    try:
        from curl_cffi import requests as cf
        r = cf.get(MINVU_PDF_URL, impersonate="chrome124", timeout=30)
        if r.status_code == 200:
            return _parsear_minvu_pdf(r.content)
    except Exception as e:
        print(f"  [MINVU] error: {e}")

    data = _descargar_minvu_playwright(MINVU_PDF_URL)
    if data:
        return _parsear_minvu_pdf(data)

    print("[MINVU] No se pudo acceder. Descarga el archivo manualmente:")
    print(f"  {MINVU_XLSX_URL}")
    print("  Luego: python scraper_precios.py --parsear-xlsx archivo.xlsx --fuente minvu")
    return []


# ─── CYPE Generador de Precios Chile ─────────────────────────────────────────

CYPE_BASE = "https://chile.generadordeprecios.info/obra_nueva"

# Categorias principales del generador CYPE Chile
CYPE_CATEGORIAS = [
    "Acondicionamiento_del_terreno.html",
    "Cimentaciones.html",
    "Estructuras.html",
    "Fachadas_y_particiones.html",
    "Carpinteria_cerrajeria_vidrios_y_protecciones_solares.html",
    "Revestimientos_y_trasdosados.html",
    "Senalizacion_y_equipamiento.html",
    "Instalaciones.html",
    "Urbanizacion_obras_exteriores_y_paisajismo.html",
    "Gestion_de_residuos.html",
    "Control_de_calidad_y_ensayos.html",
    "Seguridad_y_salud.html",
]


def _cype_extraer_items(page) -> list[dict]:
    """Extrae items de precio de una pagina CYPE."""
    partidas = []
    try:
        # CYPE usa tabla con clase 'table-prices' o similar
        filas = page.query_selector_all("tr.DataGridRow, tr.DataGridAlt, .row-price")
        for fila in filas:
            celdas = [c.inner_text().strip() for c in fila.query_selector_all("td")]
            if len(celdas) < 3:
                continue
            # Columnas tipicas: codigo | descripcion | unidad | precio
            desc = celdas[1] if len(celdas) > 1 else celdas[0]
            unidad = celdas[-2] if len(celdas) > 2 else "m2"
            precio_str = celdas[-1]
            precio = _clp(precio_str)
            if precio < 100:
                continue
            partidas.append({
                "partida": _slug(desc),
                "unidad": unidad.lower(),
                "precio_clp": precio,
                "fuente": "CYPE Chile 2025",
                "fecha": "2025",
                "_descripcion_original": desc,
            })
    except Exception as e:
        print(f"  [CYPE] error extrayendo items: {e}")
    return partidas


def scrape_cype() -> list[dict]:
    print("[CYPE] Iniciando scraping con Playwright...")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("[CYPE] Playwright no instalado:")
        print("  pip install playwright && python -m playwright install chromium")
        return []

    partidas = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="es-CL",
        )
        page = ctx.new_page()

        # Visitar home para establecer sesion
        try:
            page.goto(CYPE_BASE + "/", timeout=30000)
            time.sleep(2)
        except Exception as e:
            print(f"  [CYPE] no se pudo acceder al home: {e}")
            browser.close()
            return []

        for cat in CYPE_CATEGORIAS:
            url = f"{CYPE_BASE}/{cat}"
            try:
                page.goto(url, timeout=20000)
                time.sleep(1.5)
                nuevas = _cype_extraer_items(page)
                if nuevas:
                    print(f"  [CYPE] {cat[:40]}: {len(nuevas)} items")
                    partidas.extend(nuevas)

                # Navegar subcategorias si las hay
                links_sub = page.query_selector_all("a.subcategory-link, .chapter-link")
                for link in links_sub[:10]:  # max 10 subcats por categoria
                    href = link.get_attribute("href") or ""
                    if not href.startswith("http"):
                        href = CYPE_BASE + "/" + href.lstrip("/")
                    try:
                        page.goto(href, timeout=15000)
                        time.sleep(1)
                        sub_items = _cype_extraer_items(page)
                        if sub_items:
                            print(f"    {href.split('/')[-1][:40]}: {len(sub_items)} items")
                            partidas.extend(sub_items)
                    except Exception:
                        pass
                    page.go_back()
                    time.sleep(0.5)

            except Exception as e:
                print(f"  [CYPE] {cat}: {e}")

        browser.close()

    print(f"[CYPE] Total: {len(partidas)} partidas")
    return partidas


# ─── ONDAC ────────────────────────────────────────────────────────────────────

ONDAC_LOGIN = "https://portal.ondac.com/601/w3-channel.html"
ONDAC_APU   = "https://portal.ondac.com/601/w3-propertyvalue-122724.html"


def scrape_ondac() -> list[dict]:
    """
    Scraper ONDAC — requiere cuenta registrada.
    Credenciales via variables de entorno:
        ONDAC_USER=tucorreo@ejemplo.com
        ONDAC_PASS=tupassword
    O archivo .env en el directorio del proyecto.
    """
    user = os.environ.get("ONDAC_USER")
    passwd = os.environ.get("ONDAC_PASS")

    if not user or not passwd:
        # Intentar cargar desde .env
        env_path = Path(__file__).parent / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                if line.startswith("ONDAC_USER="):
                    user = line.split("=", 1)[1].strip()
                if line.startswith("ONDAC_PASS="):
                    passwd = line.split("=", 1)[1].strip()

    if not user or not passwd:
        print("[ONDAC] Credenciales requeridas. Crea un archivo .env:")
        print("  ONDAC_USER=tucorreo@ondac.com")
        print("  ONDAC_PASS=tupassword")
        return []

    print(f"[ONDAC] Iniciando con usuario {user[:3]}***")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("[ONDAC] Playwright no instalado: pip install playwright && python -m playwright install chromium")
        return []

    partidas = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # visible para resolver CAPTCHA si aparece
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        # Login
        try:
            page.goto(ONDAC_LOGIN, timeout=30000)
            time.sleep(2)

            # Buscar campos de login (ONDAC usa login de gobierno/clave unica o propio)
            selectors_user = ["input[type='email']", "input[name='user']", "#username", "#email"]
            selectors_pass = ["input[type='password']", "#password", "#pass"]

            for sel in selectors_user:
                el = page.query_selector(sel)
                if el:
                    el.fill(user)
                    break

            for sel in selectors_pass:
                el = page.query_selector(sel)
                if el:
                    el.fill(passwd)
                    break

            # Submit
            btn = (page.query_selector("button[type='submit']") or
                   page.query_selector("input[type='submit']") or
                   page.query_selector(".btn-login"))
            if btn:
                btn.click()
                time.sleep(3)
            else:
                page.keyboard.press("Enter")
                time.sleep(3)

            print(f"  [ONDAC] login enviado — URL actual: {page.url[:80]}")
        except Exception as e:
            print(f"  [ONDAC] error en login: {e}")
            browser.close()
            return []

        # Navegar a seccion APU
        try:
            page.goto(ONDAC_APU, timeout=30000)
            time.sleep(2)

            # Extraer tabla de precios
            filas = page.query_selector_all("table tr, .apu-row, .price-row")
            for fila in filas:
                celdas = [c.inner_text().strip() for c in fila.query_selector_all("td, .cell")]
                if len(celdas) < 3:
                    continue
                desc = celdas[0]
                if len(desc) < 4:
                    continue
                unidad = celdas[1] if len(celdas) > 1 else "m2"
                precio = _clp(celdas[2]) if len(celdas) > 2 else 0
                if precio < 100:
                    continue
                partidas.append({
                    "partida": _slug(desc),
                    "unidad": unidad.lower(),
                    "precio_clp": precio,
                    "fuente": "ONDAC 2025",
                    "fecha": "2025",
                    "_descripcion_original": desc,
                })

            print(f"  [ONDAC] {len(partidas)} partidas extraidas")
        except Exception as e:
            print(f"  [ONDAC] error extrayendo APUs: {e}")

        # Captura pantalla para debug si hay 0 resultados
        if not partidas:
            screenshot = Path("/tmp/ondac_debug.png")
            page.screenshot(path=str(screenshot))
            print(f"  [ONDAC] 0 partidas — screenshot guardado en {screenshot}")
            print("  Revisa la imagen: puede haber CAPTCHA o estructura de UI diferente.")

        browser.close()

    return partidas


# ─── Parsear archivo local ────────────────────────────────────────────────────

def parsear_xlsx_local(path: str, fuente: str = "local") -> list[dict]:
    """Parsea un Excel descargado manualmente."""
    data = Path(path).read_bytes()
    if path.endswith(".xlsx"):
        return _parsear_minvu_xlsx(data)
    elif path.endswith(".pdf"):
        return _parsear_minvu_pdf(data)
    else:
        print(f"Formato no soportado: {path}")
        return []


# ─── Diff con precios_cl.csv ──────────────────────────────────────────────────

def diff_con_existente(staging_path: str, base_path: str = "precios_cl.csv") -> None:
    """Muestra partidas nuevas en staging que no estan en base."""
    base = Path(base_path)
    staging = Path(staging_path)

    if not base.exists():
        print(f"Base no encontrada: {base_path}")
        return
    if not staging.exists():
        print(f"Staging no encontrado: {staging_path}")
        return

    partidas_base = set()
    with open(base, encoding="utf-8") as f:
        for row in csv.DictReader(row for row in f if not row.startswith("#")):
            partidas_base.add(row["partida"])

    nuevas = []
    with open(staging, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["partida"] not in partidas_base:
                nuevas.append(row)

    if not nuevas:
        print("No hay partidas nuevas en staging.")
        return

    print(f"\n{'─'*70}")
    print(f"  {len(nuevas)} partidas NUEVAS (no estan en {base_path})")
    print(f"{'─'*70}")
    print(f"{'PARTIDA':<45} {'UNIDAD':>6} {'PRECIO CLP':>12}")
    print(f"{'─'*70}")
    for r in nuevas:
        print(f"  {r['partida'][:43]:<43} {r['unidad']:>6} {int(r['precio_clp']):>12,}")

    print(f"\nPara agregar al CSV base:")
    print(f"  cat {staging_path} >> {base_path}  # verificar primero!")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scraper de precios de construccion Chile",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--fuente", choices=["minvu", "cype", "ondac", "all"],
                        default="minvu", help="Fuente a scrapear (default: minvu)")
    parser.add_argument("--output", default="precios_staging.csv",
                        help="Archivo de salida (default: precios_staging.csv)")
    parser.add_argument("--diff", metavar="STAGING",
                        help="Mostrar partidas nuevas vs precios_cl.csv")
    parser.add_argument("--parsear-xlsx", metavar="ARCHIVO",
                        help="Parsear Excel/PDF descargado manualmente")
    parser.add_argument("--base", default="precios_cl.csv",
                        help="CSV base para --diff (default: precios_cl.csv)")
    args = parser.parse_args()

    if args.diff:
        diff_con_existente(args.diff, args.base)
        return

    if args.parsear_xlsx:
        fuente_label = f"manual_{Path(args.parsear_xlsx).stem}"
        partidas = parsear_xlsx_local(args.parsear_xlsx, fuente=fuente_label)
        if partidas:
            _guardar_csv(partidas, Path(args.output))
            diff_con_existente(args.output, args.base)
        return

    todas: list[dict] = []

    if args.fuente in ("minvu", "all"):
        todas.extend(scrape_minvu())

    if args.fuente in ("cype", "all"):
        todas.extend(scrape_cype())

    if args.fuente in ("ondac", "all"):
        todas.extend(scrape_ondac())

    if not todas:
        print("\nNo se extrajeron datos. Ver mensajes arriba.")
        print("\nAlternativa: descarga el archivo manualmente y usa:")
        print("  python scraper_precios.py --parsear-xlsx archivo.xlsx")
        sys.exit(1)

    # Deduplicar por partida (ultimo valor gana)
    dedup: dict[str, dict] = {}
    for p in todas:
        if p["precio_clp"] > 0:
            dedup[p["partida"]] = p

    partidas_final = sorted(dedup.values(), key=lambda x: x["partida"])
    _guardar_csv(partidas_final, Path(args.output))
    diff_con_existente(args.output, args.base)


if __name__ == "__main__":
    main()
