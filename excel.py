"""
Excel export: genera workbook con 4 hojas (Resumen, Recintos, Cubicacion, Trazabilidad).

Usa openpyxl. Las cantidades y subtotales se escriben como formulas =SUM() reales
para que el usuario pueda editar valores y ver el total recalculado.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
_TOTAL_FILL = PatternFill("solid", fgColor="FFE4B5")


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _BOLD
        c.fill = _HEADER_FILL


def _autofit(ws, max_widths: dict[int, int] | None = None) -> None:
    max_widths = max_widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if col in max_widths:
            ws.column_dimensions[letter].width = max_widths[col]
            continue
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(50, len(str(cell.value))))
        ws.column_dimensions[letter].width = max_len + 2


def _hoja_resumen(wb: Workbook, resultado: dict, cubicacion: dict, presupuesto: dict) -> None:
    ws = wb.create_sheet("Resumen")
    lam = resultado.get("lamina", {})
    s = presupuesto["subtotales"]
    areas = cubicacion["resumen_areas"]

    ws["A1"] = "CUBICACION + PRESUPUESTO"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    rows = [
        ("Lamina:", lam.get("titulo") or "—"),
        ("Tipo:", lam.get("tipo", "—")),
        ("Escala:", lam.get("escala") or "—"),
        ("Calidad plano:", resultado.get("calidad_plano", "—")),
        ("Generado:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Area confiable (conf ≥ 0.70):", f"{areas['confiable_m2']:.1f} m²"),
        ("Area incierta (0.50-0.69):", f"{areas['incierta_m2']:.1f} m²"),
        ("Area total:", f"{areas['total_m2']:.1f} m²"),
        ("", ""),
        ("Subtotal neto:", f"$ {s['neto']:,}".replace(",", ".")),
        (f"GG + Utilidad ({s['gg_utilidad_pct']*100:.0f}%):", f"$ {s['gg_utilidad_monto']:,}".replace(",", ".")),
        ("Subtotal + GG&U:", f"$ {s['con_gg']:,}".replace(",", ".")),
        (f"IVA ({s['iva_pct']*100:.0f}%):", f"$ {s['iva_monto']:,}".replace(",", ".")),
        ("TOTAL CON IVA:", f"$ {s['total_con_iva']:,} CLP".replace(",", ".")),
    ]

    if cubicacion.get("resumen_areas", {}).get("total_m2", 0) > 0:
        costo_m2 = s["total_con_iva"] / cubicacion["resumen_areas"]["total_m2"]
        rows.append(("Costo por m²:", f"$ {int(costo_m2):,} CLP/m²".replace(",", ".")))

    # Avisos
    sin_area = cubicacion.get("recintos_sin_area", [])
    if sin_area:
        rows.append(("", ""))
        rows.append(("⚠ Recintos sin area (omitidos):", ", ".join(sin_area)))
    comunes = cubicacion.get("comunes_omitidos", [])
    if comunes:
        rows.append(("⚠ Areas comunes no cubicadas:", f"{len(comunes)} recinto(s)"))
    faltantes = presupuesto.get("faltantes", [])
    if faltantes:
        rows.append(("⚠ Partidas sin precio:", ", ".join(faltantes)))

    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = _BOLD if k.startswith("TOTAL") else Font()
        c = ws.cell(row=i, column=2, value=v)
        if k.startswith("TOTAL"):
            c.font = Font(bold=True, size=12)
            c.fill = _TOTAL_FILL

    _autofit(ws, {1: 32, 2: 45})


def _hoja_recintos(wb: Workbook, resultado: dict, cubicacion: dict) -> None:
    ws = wb.create_sheet("Recintos")
    _set_header(ws, 1, ["Departamento", "Nombre", "Categoria", "Area m²", "Perim ml", "Confianza", "Flag"])

    # Mezclar info de cubicacion (categoria, perimetro) con recintos originales
    info_por_nombre = {c["nombre"]: c for c in cubicacion.get("clasificacion_recintos", [])}

    recintos = resultado.get("recintos", [])
    # Ordenar por depto, nombre
    recintos_sorted = sorted(
        recintos, key=lambda r: ((r.get("departamento") or "~SIN").upper(), r.get("nombre", ""))
    )

    row = 2
    depto_actual = None
    depto_start = row
    depto_total = 0.0
    for rec in recintos_sorted:
        depto = rec.get("departamento") or "SIN ASIGNAR"
        if depto != depto_actual:
            if depto_actual is not None and row > depto_start:
                # Subtotal del depto anterior
                ws.cell(row=row, column=1, value=f"Subtotal {depto_actual}").font = _BOLD
                ws.cell(row=row, column=4, value=round(depto_total, 1)).font = _BOLD
                row += 1
            depto_actual = depto
            depto_start = row
            depto_total = 0.0

        info = info_por_nombre.get(rec.get("nombre", ""), {})
        area = rec.get("area_m2")
        if area:
            depto_total += area
        conf = rec.get("confianza") or 0
        flag = "✓" if conf >= 0.70 else "⚠"

        ws.cell(row=row, column=1, value=depto)
        ws.cell(row=row, column=2, value=rec.get("nombre", ""))
        ws.cell(row=row, column=3, value=info.get("categoria", "—"))
        ws.cell(row=row, column=4, value=area if area else "—")
        ws.cell(row=row, column=5, value=info.get("perimetro_ml", 0))
        ws.cell(row=row, column=6, value=round(conf, 2))
        ws.cell(row=row, column=7, value=flag)
        row += 1

    # Subtotal final
    if depto_actual is not None:
        ws.cell(row=row, column=1, value=f"Subtotal {depto_actual}").font = _BOLD
        ws.cell(row=row, column=4, value=round(depto_total, 1)).font = _BOLD

    _autofit(ws, {1: 18, 2: 32, 3: 12, 4: 10, 5: 10, 6: 12, 7: 6})


def _hoja_cubicacion(wb: Workbook, cubicacion: dict, presupuesto: dict) -> None:
    ws = wb.create_sheet("Cubicacion")
    _set_header(ws, 1, ["Partida", "Descripcion", "Unidad", "Cantidad", "Precio unit. CLP", "Subtotal CLP", "Fuente"])

    row = 2
    primera_fila = row
    tipo_actual = None
    for l in presupuesto["lineas"]:
        tipo = l.get("tipo", "residencial")
        if tipo != tipo_actual:
            if tipo == "vial":
                row += 1
                hdr = ws.cell(row=row, column=1, value="INFRAESTRUCTURA VIAL")
                hdr.font = Font(bold=True)
                hdr.fill = PatternFill("solid", fgColor="C6EFCE")
                ws.cell(row=row, column=2, value="Cantidades estimadas — verificar con planos de seccion").font = Font(italic=True)
                row += 1
            tipo_actual = tipo
        ws.cell(row=row, column=1, value=l["partida"])
        ws.cell(row=row, column=2, value=l["descripcion"])
        ws.cell(row=row, column=3, value=l["unidad"])
        ws.cell(row=row, column=4, value=l["cantidad"])
        ws.cell(row=row, column=5, value=l["precio_unitario_clp"])
        # Subtotal como formula: =D{row}*E{row}
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=7, value=l["fuente"])
        row += 1
    ultima_fila = row - 1

    # Subtotales con formulas
    s = presupuesto["subtotales"]
    row += 1
    ws.cell(row=row, column=2, value="SUBTOTAL NETO").font = _BOLD
    ws.cell(row=row, column=6, value=f"=SUM(F{primera_fila}:F{ultima_fila})").font = _BOLD
    sub_row = row

    row += 1
    ws.cell(row=row, column=2, value=f"GG + Utilidad ({s['gg_utilidad_pct']*100:.0f}%)")
    ws.cell(row=row, column=6, value=f"=F{sub_row}*{s['gg_utilidad_pct']}")
    gg_row = row

    row += 1
    ws.cell(row=row, column=2, value="SUBTOTAL + GG&U").font = _BOLD
    ws.cell(row=row, column=6, value=f"=F{sub_row}+F{gg_row}").font = _BOLD
    con_gg_row = row

    row += 1
    ws.cell(row=row, column=2, value=f"IVA ({s['iva_pct']*100:.0f}%)")
    ws.cell(row=row, column=6, value=f"=F{con_gg_row}*{s['iva_pct']}")
    iva_row = row

    row += 1
    total_cell = ws.cell(row=row, column=2, value="TOTAL CON IVA")
    total_cell.font = Font(bold=True, size=12)
    final = ws.cell(row=row, column=6, value=f"=F{con_gg_row}+F{iva_row}")
    final.font = Font(bold=True, size=12)
    final.fill = _TOTAL_FILL

    # Areas comunes no cubicadas (informativo)
    comunes = cubicacion.get("comunes_omitidos", [])
    if comunes:
        row += 3
        ws.cell(row=row, column=1, value="AREAS COMUNES (no cubicadas)").font = _BOLD
        row += 1
        _set_header(ws, row, ["Nombre", "Area m²"])
        row += 1
        for c in comunes:
            ws.cell(row=row, column=1, value=c["nombre"])
            ws.cell(row=row, column=2, value=c["area_m2"])
            row += 1

    _autofit(ws, {1: 36, 2: 38, 3: 8, 4: 10, 5: 16, 6: 18, 7: 18})


def _hoja_trazabilidad(wb: Workbook, resultado: dict, schedule: dict | None, cubicacion: dict) -> None:
    ws = wb.create_sheet("Trazabilidad")
    row = 1

    ws.cell(row=row, column=1, value="SCHEDULE EXTRAIDO").font = Font(bold=True, size=12)
    row += 2
    if schedule and schedule.get("tiene_tabla"):
        _set_header(ws, row, ["Departamento", "Area total m²", "Recintos en tabla"])
        row += 1
        for d in schedule.get("departamentos", []):
            recs_str = "; ".join(
                f"{r['nombre']}: {r.get('area_m2', '?')}m²" for r in d.get("recintos", [])
            )
            ws.cell(row=row, column=1, value=d.get("id", "?"))
            ws.cell(row=row, column=2, value=d.get("area_total_m2", "—"))
            ws.cell(row=row, column=3, value=recs_str)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(sin tabla de areas detectada)")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="OBSERVACIONES API").font = Font(bold=True, size=12)
    row += 1
    obs = resultado.get("observaciones", "")
    if obs:
        for line in obs.split(" | "):
            ws.cell(row=row, column=1, value=line.strip()[:200])
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="COSTO Y TOKENS API").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Tiles procesados:")
    ws.cell(row=row, column=2, value=resultado.get("_tiles_procesados", 1))
    row += 1
    ws.cell(row=row, column=1, value="Tokens entrada:")
    ws.cell(row=row, column=2, value=resultado.get("_tokens_entrada", 0))
    row += 1
    ws.cell(row=row, column=1, value="Tokens salida:")
    ws.cell(row=row, column=2, value=resultado.get("_tokens_salida", 0))
    row += 1
    ti = resultado.get("_tokens_entrada", 0)
    to = resultado.get("_tokens_salida", 0)
    costo_usd = (ti * 15 + to * 75) / 1_000_000
    ws.cell(row=row, column=1, value="Costo API USD:")
    ws.cell(row=row, column=2, value=f"${costo_usd:.4f}")

    # Vanos con dimension estimada
    row += 3
    estimados = [v for v in cubicacion.get("vanos_detalle", []) if v.get("dimension_estimada")]
    if estimados:
        ws.cell(row=row, column=1, value="VANOS CON DIMENSION ESTIMADA").font = Font(bold=True, size=12)
        row += 1
        _set_header(ws, row, ["Tipo", "Cantidad", "Ancho × Alto", "Area unit.", "Area total"])
        row += 1
        for v in estimados:
            ws.cell(row=row, column=1, value=v["tipo"])
            ws.cell(row=row, column=2, value=v["cantidad"])
            ws.cell(row=row, column=3, value=f"{v['ancho_m']:.2f} × {v['alto_m']:.2f}")
            ws.cell(row=row, column=4, value=v["area_unitaria_m2"])
            ws.cell(row=row, column=5, value=v["area_total_m2"])
            row += 1

    # Recintos con clasificacion heuristica
    row += 2
    heuristicos = [c for c in cubicacion.get("clasificacion_recintos", []) if c.get("clasificacion_heuristica")]
    if heuristicos:
        ws.cell(row=row, column=1, value="RECINTOS CON CLASIFICACION HEURISTICA").font = Font(bold=True, size=12)
        row += 1
        _set_header(ws, row, ["Nombre", "Categoria asignada"])
        row += 1
        for c in heuristicos:
            ws.cell(row=row, column=1, value=c["nombre"])
            ws.cell(row=row, column=2, value=c["categoria"])
            row += 1

    _autofit(ws, {1: 36, 2: 24, 3: 28, 4: 14, 5: 14})


def exportar_excel(
    resultado: dict,
    schedule: dict | None,
    cubicacion: dict,
    presupuesto: dict,
    ruta_salida: str | Path,
) -> Path:
    """Genera el workbook completo y lo guarda en ruta_salida. Devuelve Path."""
    wb = Workbook()
    # Eliminar la hoja por defecto que crea openpyxl
    default = wb.active
    wb.remove(default)

    _hoja_resumen(wb, resultado, cubicacion, presupuesto)
    _hoja_recintos(wb, resultado, cubicacion)
    _hoja_cubicacion(wb, cubicacion, presupuesto)
    _hoja_trazabilidad(wb, resultado, schedule, cubicacion)

    ruta = Path(ruta_salida)
    wb.save(ruta)
    return ruta
